"""Grand Aceh Kuliner POS - backend regression suite."""
import io
import uuid

import pytest
import requests

from conftest import API

UQ = uuid.uuid4().hex[:6]


def _tag(s):
    return f"TEST_{s}_{UQ}"


# ---------------------------------------------------------------- AUTH
class TestAuth:
    def test_admin_login(self, admin_creds):
        r = requests.post(f"{API}/auth/login", json=admin_creds, timeout=30)
        assert r.status_code == 200, r.text
        d = r.json()
        assert isinstance(d["token"], str) and len(d["token"]) > 10
        assert d["user"]["role"] == "admin"
        assert d["user"]["email"] == admin_creds["email"].lower()

    def test_kasir_login(self, kasir_creds):
        r = requests.post(f"{API}/auth/login", json=kasir_creds, timeout=30)
        assert r.status_code == 200, r.text
        assert r.json()["user"]["role"] == "kasir"

    def test_login_bad_password(self, admin_creds):
        r = requests.post(f"{API}/auth/login",
                          json={"email": admin_creds["email"], "password": "wrong-pass-xyz"}, timeout=30)
        assert r.status_code == 401

    def test_me(self, admin):
        r = admin.get(f"{API}/auth/me", timeout=30)
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["role"] == "admin"
        assert "password_hash" not in d
        assert "_id" not in d

    def test_me_no_token(self, anon):
        assert anon.get(f"{API}/auth/me", timeout=30).status_code == 401

    def test_me_invalid_token(self, anon):
        r = anon.get(f"{API}/auth/me", headers={"Authorization": "Bearer garbage.token.xx"}, timeout=30)
        assert r.status_code == 401

    def test_bcrypt_hash_format(self):
        import asyncio, os
        from motor.motor_asyncio import AsyncIOMotorClient
        from dotenv import dotenv_values
        env = dotenv_values("/app/backend/.env")

        async def go():
            c = AsyncIOMotorClient(env["MONGO_URL"])
            u = await c[env["DB_NAME"]].users.find_one({"role": "admin"})
            c.close()
            return u
        u = asyncio.run(go())
        assert u is not None
        assert u["password_hash"].startswith("$2b$"), u["password_hash"][:6]

    def test_kasir_cannot_list_users(self, kasir):
        assert kasir.get(f"{API}/users", timeout=30).status_code == 403


# ---------------------------------------------------------------- CATEGORIES
class TestCategories:
    created = []

    def test_list_categories(self, admin):
        r = admin.get(f"{API}/categories", timeout=30)
        assert r.status_code == 200
        cats = r.json()
        assert len(cats) > 0
        for c in cats:
            assert "_id" not in c
            assert c["type"] in ("makanan", "minuman", "retail")

    def test_create_update_category(self, admin):
        r = admin.post(f"{API}/categories", json={"name": _tag("KAT"), "type": "retail", "sort_order": 99}, timeout=30)
        assert r.status_code == 200, r.text
        c = r.json()
        TestCategories.created.append(c["id"])
        assert c["type"] == "retail"
        r2 = admin.put(f"{API}/categories/{c['id']}",
                       json={"name": _tag("KAT2"), "type": "makanan", "sort_order": 5, "active": True}, timeout=30)
        assert r2.status_code == 200, r2.text
        assert r2.json()["type"] == "makanan"
        # verify persistence
        got = [x for x in admin.get(f"{API}/categories", timeout=30).json() if x["id"] == c["id"]][0]
        assert got["name"] == _tag("KAT2")

    def test_kasir_cannot_create_category(self, kasir):
        r = kasir.post(f"{API}/categories", json={"name": "TEST_forbidden", "type": "retail"}, timeout=30)
        assert r.status_code == 403

    def test_update_nonexistent_category(self, admin):
        r = admin.put(f"{API}/categories/no-such-id", json={"name": "x", "type": "retail"}, timeout=30)
        assert r.status_code == 404

    def test_delete_unused_category_hard_deletes(self, admin):
        c = admin.post(f"{API}/categories", json={"name": _tag("KATDEL"), "type": "retail"}, timeout=30).json()
        r = admin.delete(f"{API}/categories/{c['id']}", timeout=30)
        assert r.status_code == 200 and r.json().get("deleted") is True
        assert not any(x["id"] == c["id"] for x in admin.get(f"{API}/categories", timeout=30).json())

    def test_delete_used_category_soft_deactivates(self, admin):
        c = admin.post(f"{API}/categories", json={"name": _tag("KATUSED"), "type": "retail"}, timeout=30).json()
        p = admin.post(f"{API}/products", json={"name": _tag("P_KATUSED"), "sku": _tag("SKUKU"),
                                                "category_id": c["id"], "type": "retail", "price": 1000,
                                                "stock": 5}, timeout=30)
        assert p.status_code == 200, p.text
        r = admin.delete(f"{API}/categories/{c['id']}", timeout=30)
        assert r.status_code == 200, r.text
        assert r.json().get("soft_deleted") is True
        got = [x for x in admin.get(f"{API}/categories", timeout=30).json() if x["id"] == c["id"]]
        assert got and got[0]["active"] is False
        admin.delete(f"{API}/products/{p.json()['id']}", timeout=30)
        TestCategories.created.append(c["id"])


# ---------------------------------------------------------------- PRODUCTS
@pytest.fixture(scope="module")
def cat_map(admin):
    cats = admin.get(f"{API}/categories", timeout=30).json()
    out = {}
    for c in cats:
        if c.get("active", True) and not c["name"].startswith("TEST_"):
            out.setdefault(c["type"], c)
    assert {"makanan", "minuman", "retail"} <= set(out), f"missing seeded categories: {list(out)}"
    return out


@pytest.fixture(scope="module")
def prod_map(admin):
    prods = admin.get(f"{API}/products", timeout=30).json()
    out = {}
    for p in prods:
        if p.get("active", True) and not p.get("sold_out") and not p["name"].startswith("TEST_"):
            out.setdefault(p["type"], p)
    assert {"makanan", "minuman", "retail"} <= set(out), f"missing seeded products: {list(out)}"
    return out


class TestProducts:
    def test_seed_products_present(self, admin):
        prods = admin.get(f"{API}/products", timeout=30).json()
        assert len(prods) >= 13
        for p in prods:
            assert "_id" not in p
        retail = [p for p in prods if p["type"] == "retail"]
        fnb = [p for p in prods if p["type"] in ("makanan", "minuman")]
        assert retail and fnb
        assert all(p.get("track_stock") for p in retail), "retail products must track stock"
        assert all(not p.get("track_stock") for p in fnb), "F&B must not track stock"

    def test_filter_by_type(self, admin):
        r = admin.get(f"{API}/products", params={"type": "retail"}, timeout=30)
        assert r.status_code == 200
        assert all(p["type"] == "retail" for p in r.json())

    def test_create_duplicate_sku_rejected(self, admin, cat_map):
        sku = _tag("SKUDUP")
        base = {"name": _tag("P1"), "sku": sku, "category_id": cat_map["retail"]["id"],
                "type": "retail", "price": 5000, "stock": 3}
        r1 = admin.post(f"{API}/products", json=base, timeout=30)
        assert r1.status_code == 200, r1.text
        pid = r1.json()["id"]
        r2 = admin.post(f"{API}/products", json={**base, "name": _tag("P2")}, timeout=30)
        assert r2.status_code == 400, r2.text
        assert "SKU" in r2.json()["detail"]
        admin.delete(f"{API}/products/{pid}", timeout=30)

    def test_negative_price_rejected(self, admin, cat_map):
        r = admin.post(f"{API}/products", json={"name": _tag("PNEG"), "sku": _tag("SKUNEG"),
                                                "category_id": cat_map["makanan"]["id"],
                                                "type": "makanan", "price": -1}, timeout=30)
        assert r.status_code == 400

    def test_invalid_category_rejected(self, admin):
        r = admin.post(f"{API}/products", json={"name": _tag("PBADCAT"), "sku": _tag("SKUBC"),
                                                "category_id": "not-a-real-id", "type": "makanan",
                                                "price": 1000}, timeout=30)
        assert r.status_code == 400

    def test_invalid_type_rejected(self, admin, cat_map):
        r = admin.post(f"{API}/products", json={"name": _tag("PBT"), "sku": _tag("SKUBT"),
                                                "category_id": cat_map["makanan"]["id"],
                                                "type": "elektronik", "price": 1000}, timeout=30)
        assert r.status_code == 422

    def test_update_and_sold_out_toggle(self, admin, cat_map):
        r = admin.post(f"{API}/products", json={"name": _tag("PUPD"), "sku": _tag("SKUUPD"),
                                                "category_id": cat_map["minuman"]["id"],
                                                "type": "minuman", "price": 8000}, timeout=30)
        assert r.status_code == 200, r.text
        pid = r.json()["id"]
        up = admin.put(f"{API}/products/{pid}", json={"name": _tag("PUPD2"), "sku": _tag("SKUUPD"),
                                                      "category_id": cat_map["minuman"]["id"],
                                                      "type": "minuman", "price": 9500}, timeout=30)
        assert up.status_code == 200, up.text
        assert up.json()["price"] == 9500
        t1 = admin.patch(f"{API}/products/{pid}/sold-out", timeout=30)
        assert t1.status_code == 200 and t1.json()["sold_out"] is True
        got = [p for p in admin.get(f"{API}/products", timeout=30).json() if p["id"] == pid][0]
        assert got["sold_out"] is True
        t2 = admin.patch(f"{API}/products/{pid}/sold-out", timeout=30)
        assert t2.json()["sold_out"] is False
        admin.delete(f"{API}/products/{pid}", timeout=30)

    def test_delete_unused_product_hard_delete(self, admin, cat_map):
        pid = admin.post(f"{API}/products", json={"name": _tag("PDEL"), "sku": _tag("SKUDEL"),
                                                  "category_id": cat_map["retail"]["id"],
                                                  "type": "retail", "price": 3000, "stock": 1},
                         timeout=30).json()["id"]
        r = admin.delete(f"{API}/products/{pid}", timeout=30)
        assert r.json().get("deleted") is True
        assert not any(p["id"] == pid for p in admin.get(f"{API}/products", timeout=30).json())

    def test_delete_used_product_soft_deactivates(self, admin, cat_map, pm_cash):
        p = admin.post(f"{API}/products", json={"name": _tag("PUSED"), "sku": _tag("SKUUSED"),
                                                "category_id": cat_map["retail"]["id"],
                                                "type": "retail", "price": 4000, "stock": 10},
                       timeout=30).json()
        o = admin.post(f"{API}/orders", json={"order_type": "retail", "items": [
            {"product_id": p["id"], "name": p["name"], "price": 4000, "qty": 1, "type": "retail"}]}, timeout=30)
        assert o.status_code == 200, o.text
        r = admin.delete(f"{API}/products/{p['id']}", timeout=30)
        assert r.status_code == 200 and r.json().get("soft_deleted") is True
        got = [x for x in admin.get(f"{API}/products", timeout=30).json() if x["id"] == p["id"]][0]
        assert got["active"] is False

    def test_kasir_cannot_create_product(self, kasir, cat_map):
        r = kasir.post(f"{API}/products", json={"name": "TEST_x", "sku": _tag("SKUK"),
                                                "category_id": cat_map["retail"]["id"],
                                                "type": "retail", "price": 100}, timeout=30)
        assert r.status_code == 403


# ---------------------------------------------------------------- TABLES
class TestTables:
    def test_list_tables(self, admin):
        r = admin.get(f"{API}/tables", timeout=30)
        assert r.status_code == 200
        tables = r.json()
        assert len(tables) >= 8
        for t in tables:
            assert "_id" not in t
            assert t["status"] in ("empty", "open_bill")

    def test_duplicate_name_rejected(self, admin):
        name = _tag("MEJA")
        r1 = admin.post(f"{API}/tables", json={"name": name, "area": "TEST", "capacity": 2}, timeout=30)
        assert r1.status_code == 200, r1.text
        tid = r1.json()["id"]
        r2 = admin.post(f"{API}/tables", json={"name": name, "area": "TEST", "capacity": 4}, timeout=30)
        assert r2.status_code == 400
        admin.delete(f"{API}/tables/{tid}", timeout=30)

    def test_kasir_cannot_create_table(self, kasir):
        assert kasir.post(f"{API}/tables", json={"name": _tag("MJK")}, timeout=30).status_code == 403

    def test_open_bill_blocks_deactivate_and_delete(self, admin, prod_map):
        t = admin.post(f"{API}/tables", json={"name": _tag("MEJAOB"), "area": "TEST", "capacity": 2},
                       timeout=30).json()
        fnb = prod_map["makanan"]
        o = admin.post(f"{API}/orders", json={"order_type": "dine_in", "table_id": t["id"], "items": [
            {"product_id": fnb["id"], "name": fnb["name"], "price": fnb["price"], "qty": 1, "type": "makanan"}]},
            timeout=30)
        assert o.status_code == 200, o.text
        oid = o.json()["id"]
        # table shows open_bill
        tt = [x for x in admin.get(f"{API}/tables", timeout=30).json() if x["id"] == t["id"]][0]
        assert tt["status"] == "open_bill" and tt["open_order_id"] == oid
        # cannot deactivate
        d = admin.put(f"{API}/tables/{t['id']}", json={"name": t["name"], "area": "TEST",
                                                       "capacity": 2, "active": False}, timeout=30)
        assert d.status_code == 400, d.text
        # cannot delete
        dd = admin.delete(f"{API}/tables/{t['id']}", timeout=30)
        assert dd.status_code == 400, dd.text
        # pay it off then delete -> soft deactivate (used in transaction)
        pm = admin.get(f"{API}/payment-methods", timeout=30).json()
        cash = [p for p in pm if p["type"] == "cash"][0]
        pay = admin.post(f"{API}/orders/{oid}/pay", json={"payment_method": cash["id"],
                                                          "amount_paid": 1000000}, timeout=30)
        assert pay.status_code == 200, pay.text
        r = admin.delete(f"{API}/tables/{t['id']}", timeout=30)
        assert r.status_code == 200 and r.json().get("soft_deleted") is True
        got = [x for x in admin.get(f"{API}/tables", timeout=30).json() if x["id"] == t["id"]][0]
        assert got["active"] is False


# ---------------------------------------------------------------- PAYMENT METHODS
@pytest.fixture(scope="module")
def pm_cash(admin):
    pms = admin.get(f"{API}/payment-methods", timeout=30).json()
    cash = [p for p in pms if p["type"] == "cash" and p.get("active")]
    assert cash, "No active cash payment method seeded"
    return cash[0]


class TestPaymentMethods:
    def test_seeded_three_methods(self, admin):
        pms = admin.get(f"{API}/payment-methods", timeout=30).json()
        types = {p["type"] for p in pms}
        assert {"cash", "qris", "card"} <= types
        for p in pms:
            assert "_id" not in p


# ---------------------------------------------------------------- ORDER RULES
class TestOrderRules:
    def test_retail_order_with_fnb_item_rejected(self, admin, prod_map):
        p = prod_map["makanan"]
        r = admin.post(f"{API}/orders", json={"order_type": "retail", "items": [
            {"product_id": p["id"], "name": p["name"], "price": p["price"], "qty": 1, "type": "makanan"}]},
            timeout=30)
        assert r.status_code == 400, r.text
        assert "retail" in r.json()["detail"].lower()

    def test_take_away_with_retail_item_rejected(self, admin, prod_map):
        p = prod_map["retail"]
        r = admin.post(f"{API}/orders", json={"order_type": "take_away", "items": [
            {"product_id": p["id"], "name": p["name"], "price": p["price"], "qty": 1, "type": "retail"}]},
            timeout=30)
        assert r.status_code == 400, r.text

    def test_dine_in_with_retail_item_rejected(self, admin, prod_map):
        tables = [t for t in admin.get(f"{API}/tables", timeout=30).json()
                  if t["status"] == "empty" and t.get("active", True)]
        assert tables
        p = prod_map["retail"]
        r = admin.post(f"{API}/orders", json={"order_type": "dine_in", "table_id": tables[0]["id"], "items": [
            {"product_id": p["id"], "name": p["name"], "price": p["price"], "qty": 1, "type": "retail"}]},
            timeout=30)
        assert r.status_code == 400, r.text

    def test_dine_in_without_table_rejected(self, admin, prod_map):
        p = prod_map["makanan"]
        r = admin.post(f"{API}/orders", json={"order_type": "dine_in", "items": [
            {"product_id": p["id"], "name": p["name"], "price": p["price"], "qty": 1, "type": "makanan"}]},
            timeout=30)
        assert r.status_code == 400, r.text

    def test_take_away_with_table_rejected(self, admin, prod_map):
        tables = admin.get(f"{API}/tables", timeout=30).json()
        p = prod_map["minuman"]
        r = admin.post(f"{API}/orders", json={"order_type": "take_away", "table_id": tables[0]["id"], "items": [
            {"product_id": p["id"], "name": p["name"], "price": p["price"], "qty": 1, "type": "minuman"}]},
            timeout=30)
        assert r.status_code == 400, r.text

    def test_empty_cart_rejected(self, admin):
        r = admin.post(f"{API}/orders", json={"order_type": "take_away", "items": []}, timeout=30)
        assert r.status_code == 400


# ---------------------------------------------------------------- PAYMENT / DISCOUNT
class TestPayFlow:
    def test_take_away_pay_with_percent_discount(self, admin, prod_map, pm_cash):
        p = prod_map["makanan"]
        items = [{"product_id": p["id"], "name": p["name"], "price": p["price"], "qty": 2, "type": p["type"]}]
        o = admin.post(f"{API}/orders", json={"order_type": "take_away", "items": items}, timeout=30)
        assert o.status_code == 200, o.text
        order = o.json()
        assert order["status"] == "open"
        assert order["order_number"].startswith("GAK-")
        subtotal = p["price"] * 2
        assert order["subtotal"] == subtotal
        expected_total = round(subtotal - round(subtotal * 0.1, 2), 2)
        paid = admin.post(f"{API}/orders/{order['id']}/pay",
                          json={"payment_method": pm_cash["id"], "discount_type": "percent",
                                "discount_value": 10, "amount_paid": subtotal}, timeout=30)
        assert paid.status_code == 200, paid.text
        d = paid.json()
        assert d["status"] == "paid"
        assert d["discount"] == round(subtotal * 0.1, 2)
        assert d["total"] == expected_total
        assert d["change"] == round(subtotal - expected_total, 2)
        assert d["payment_method_name"] == pm_cash["name"]
        # persistence
        g = admin.get(f"{API}/orders/{order['id']}", timeout=30).json()
        assert g["status"] == "paid" and g["total"] == expected_total

    def test_amount_discount_capped_at_subtotal(self, admin, prod_map, pm_cash):
        p = prod_map["minuman"]
        items = [{"product_id": p["id"], "name": p["name"], "price": p["price"], "qty": 1, "type": p["type"]}]
        o = admin.post(f"{API}/orders", json={"order_type": "take_away", "items": items}, timeout=30).json()
        paid = admin.post(f"{API}/orders/{o['id']}/pay",
                          json={"payment_method": pm_cash["id"], "discount_type": "amount",
                                "discount_value": 99999999, "amount_paid": 0}, timeout=30)
        assert paid.status_code == 200, paid.text
        assert paid.json()["discount"] == p["price"]
        assert paid.json()["total"] == 0

    def test_paid_order_cannot_be_paid_again(self, admin, prod_map, pm_cash):
        p = prod_map["makanan"]
        items = [{"product_id": p["id"], "name": p["name"], "price": p["price"], "qty": 1, "type": p["type"]}]
        o = admin.post(f"{API}/orders", json={"order_type": "take_away", "items": items,
                                              "pay_now": True, "payment_method": pm_cash["id"]}, timeout=30)
        assert o.status_code == 200, o.text
        assert o.json()["status"] == "paid"
        again = admin.post(f"{API}/orders/{o.json()['id']}/pay",
                           json={"payment_method": pm_cash["id"]}, timeout=30)
        assert again.status_code == 400

    def test_paid_order_items_cannot_be_edited(self, admin, prod_map, pm_cash):
        p = prod_map["makanan"]
        items = [{"product_id": p["id"], "name": p["name"], "price": p["price"], "qty": 1, "type": p["type"]}]
        o = admin.post(f"{API}/orders", json={"order_type": "take_away", "items": items,
                                              "pay_now": True, "payment_method": pm_cash["id"]}, timeout=30).json()
        r = admin.patch(f"{API}/orders/{o['id']}/items", json={"items": items}, timeout=30)
        assert r.status_code == 400, r.text

    def test_invalid_payment_method_rejected(self, admin, prod_map):
        p = prod_map["minuman"]
        items = [{"product_id": p["id"], "name": p["name"], "price": p["price"], "qty": 1, "type": p["type"]}]
        o = admin.post(f"{API}/orders", json={"order_type": "take_away", "items": items}, timeout=30).json()
        r = admin.post(f"{API}/orders/{o['id']}/pay", json={"payment_method": "bogus-pm"}, timeout=30)
        assert r.status_code == 400

    def test_retail_stock_decrements_after_paid(self, admin, cat_map, pm_cash):
        p = admin.post(f"{API}/products", json={"name": _tag("PSTK"), "sku": _tag("SKUSTK"),
                                                "category_id": cat_map["retail"]["id"],
                                                "type": "retail", "price": 5000, "stock": 20},
                       timeout=30).json()
        items = [{"product_id": p["id"], "name": p["name"], "price": 5000, "qty": 3, "type": "retail"}]
        o = admin.post(f"{API}/orders", json={"order_type": "retail", "items": items,
                                              "pay_now": True, "payment_method": pm_cash["id"]}, timeout=30)
        assert o.status_code == 200, o.text
        got = [x for x in admin.get(f"{API}/products", timeout=30).json() if x["id"] == p["id"]][0]
        assert got["stock"] == 17, got

    def test_get_nonexistent_order_404(self, admin):
        assert admin.get(f"{API}/orders/nope-id", timeout=30).status_code == 404


# ---------------------------------------------------------------- OPEN BILL
class TestOpenBill:
    def test_dine_in_open_bill_append_and_pay(self, admin, prod_map, pm_cash):
        t = admin.post(f"{API}/tables", json={"name": _tag("MEJAOP"), "area": "TEST", "capacity": 4},
                       timeout=30).json()
        f1, f2 = prod_map["makanan"], prod_map["minuman"]
        items = [{"product_id": f1["id"], "name": f1["name"], "price": f1["price"], "qty": 1, "type": f1["type"]}]
        o = admin.post(f"{API}/orders", json={"order_type": "dine_in", "table_id": t["id"],
                                              "items": items}, timeout=30)
        assert o.status_code == 200, o.text
        order = o.json()
        assert order["status"] == "open"
        tt = [x for x in admin.get(f"{API}/tables", timeout=30).json() if x["id"] == t["id"]][0]
        assert tt["status"] == "open_bill"
        # append
        items2 = items + [{"product_id": f2["id"], "name": f2["name"], "price": f2["price"],
                           "qty": 2, "type": f2["type"]}]
        up = admin.patch(f"{API}/orders/{order['id']}/items", json={"items": items2}, timeout=30)
        assert up.status_code == 200, up.text
        expected = f1["price"] + f2["price"] * 2
        assert up.json()["subtotal"] == expected
        assert up.json()["total"] == expected
        assert len(up.json()["items"]) == 2
        # appending a retail item must fail
        rp = prod_map["retail"]
        bad = admin.patch(f"{API}/orders/{order['id']}/items", json={"items": items2 + [
            {"product_id": rp["id"], "name": rp["name"], "price": rp["price"], "qty": 1, "type": "retail"}]},
            timeout=30)
        assert bad.status_code == 400, bad.text
        # pay closes bill
        pay = admin.post(f"{API}/orders/{order['id']}/pay",
                         json={"payment_method": pm_cash["id"], "amount_paid": expected}, timeout=30)
        assert pay.status_code == 200, pay.text
        assert pay.json()["status"] == "paid"
        tt2 = [x for x in admin.get(f"{API}/tables", timeout=30).json() if x["id"] == t["id"]][0]
        assert tt2["status"] == "empty"


# ---------------------------------------------------------------- VOID / REFUND
class TestVoidRefund:
    def test_kasir_cannot_void(self, admin, kasir, prod_map, pm_cash):
        p = prod_map["makanan"]
        items = [{"product_id": p["id"], "name": p["name"], "price": p["price"], "qty": 1, "type": p["type"]}]
        o = admin.post(f"{API}/orders", json={"order_type": "take_away", "items": items,
                                              "pay_now": True, "payment_method": pm_cash["id"]}, timeout=30).json()
        r = kasir.post(f"{API}/orders/{o['id']}/void", json={"reason": "TEST", "action": "void"}, timeout=30)
        assert r.status_code == 403, r.text

    def test_void_restores_retail_stock_and_audits(self, admin, cat_map, pm_cash):
        p = admin.post(f"{API}/products", json={"name": _tag("PVOID"), "sku": _tag("SKUVOID"),
                                                "category_id": cat_map["retail"]["id"],
                                                "type": "retail", "price": 7000, "stock": 10},
                       timeout=30).json()
        items = [{"product_id": p["id"], "name": p["name"], "price": 7000, "qty": 4, "type": "retail"}]
        o = admin.post(f"{API}/orders", json={"order_type": "retail", "items": items,
                                              "pay_now": True, "payment_method": pm_cash["id"]}, timeout=30).json()
        mid = [x for x in admin.get(f"{API}/products", timeout=30).json() if x["id"] == p["id"]][0]
        assert mid["stock"] == 6
        v = admin.post(f"{API}/orders/{o['id']}/void", json={"reason": "TEST_void_reason", "action": "void"},
                       timeout=30)
        assert v.status_code == 200, v.text
        assert v.json()["status"] == "void"
        back = [x for x in admin.get(f"{API}/products", timeout=30).json() if x["id"] == p["id"]][0]
        assert back["stock"] == 10, back
        logs = admin.get(f"{API}/audit-logs", timeout=30)
        assert logs.status_code == 200
        entry = [l for l in logs.json() if l["order_id"] == o["id"]]
        assert entry, "no audit log written"
        assert entry[0]["reason"] == "TEST_void_reason"
        assert entry[0]["action"] == "void"
        assert "_id" not in entry[0]
        # double void rejected
        assert admin.post(f"{API}/orders/{o['id']}/void", json={"reason": "x"}, timeout=30).status_code == 400

    def test_refund_action(self, admin, prod_map, pm_cash):
        p = prod_map["minuman"]
        items = [{"product_id": p["id"], "name": p["name"], "price": p["price"], "qty": 1, "type": p["type"]}]
        o = admin.post(f"{API}/orders", json={"order_type": "take_away", "items": items,
                                              "pay_now": True, "payment_method": pm_cash["id"]}, timeout=30).json()
        r = admin.post(f"{API}/orders/{o['id']}/refund" if False else f"{API}/orders/{o['id']}/void",
                       json={"reason": "TEST_refund", "action": "refund"}, timeout=30)
        assert r.status_code == 200, r.text
        assert r.json()["status"] == "refunded"
        assert admin.get(f"{API}/orders/{o['id']}", timeout=30).json()["status"] == "refunded"

    def test_kasir_cannot_read_audit_logs(self, kasir):
        assert kasir.get(f"{API}/audit-logs", timeout=30).status_code == 403


# ---------------------------------------------------------------- SHIFTS
class TestShifts:
    def test_shift_lifecycle(self, kasir, admin, prod_map, pm_cash):
        cur = kasir.get(f"{API}/shifts/current", timeout=30)
        assert cur.status_code == 200
        if cur.json():
            kasir.post(f"{API}/shifts/close", json={"closing_cash": 0}, timeout=30)
        op = kasir.post(f"{API}/shifts/open", json={"opening_cash": 100000}, timeout=30)
        assert op.status_code == 200, op.text
        assert op.json()["status"] == "open"
        dup = kasir.post(f"{API}/shifts/open", json={"opening_cash": 5000}, timeout=30)
        assert dup.status_code == 400, dup.text
        p = prod_map["makanan"]
        items = [{"product_id": p["id"], "name": p["name"], "price": p["price"], "qty": 1, "type": p["type"]}]
        o = kasir.post(f"{API}/orders", json={"order_type": "take_away", "items": items,
                                              "pay_now": True, "payment_method": pm_cash["id"]}, timeout=30)
        assert o.status_code == 200, o.text
        assert o.json()["shift_id"] == op.json()["id"], "paid order not attached to shift"
        cl = kasir.post(f"{API}/shifts/close", json={"closing_cash": 100000 + p["price"]}, timeout=30)
        assert cl.status_code == 200, cl.text
        rep = cl.json()["report"]
        assert set(rep["by_type"]) >= {"dine_in", "take_away", "retail"}
        assert rep["order_count"] >= 1
        assert rep["expected_cash"] == 100000 + p["price"]
        assert kasir.get(f"{API}/shifts/current", timeout=30).json() in (None, {})

    def test_close_without_open_shift(self, kasir):
        kasir.post(f"{API}/shifts/close", json={"closing_cash": 0}, timeout=30)
        r = kasir.post(f"{API}/shifts/close", json={"closing_cash": 0}, timeout=30)
        assert r.status_code == 400

    def test_kasir_cannot_list_shifts(self, kasir):
        assert kasir.get(f"{API}/shifts", timeout=30).status_code == 403

    def test_admin_list_shifts(self, admin):
        r = admin.get(f"{API}/shifts", timeout=30)
        assert r.status_code == 200 and isinstance(r.json(), list)


# ---------------------------------------------------------------- REPORTS
class TestReports:
    def test_summary_today(self, admin):
        r = admin.get(f"{API}/reports/summary", params={"date": __import__("datetime").date.today().isoformat()},
                      timeout=30)
        assert r.status_code == 200, r.text
        d = r.json()
        for k in ("by_type", "by_payment", "fnb_total", "retail_total", "top_products",
                  "total_discount", "total_sales", "order_count"):
            assert k in d, k
        assert set(d["by_type"]) == {"dine_in", "take_away", "retail"}
        assert d["by_type"]["dine_in"].keys() >= {"count", "total"}
        assert round(d["fnb_total"] + d["retail_total"], 2) == round(d["total_sales"], 2)

    def test_summary_no_date_defaults_today(self, admin):
        r = admin.get(f"{API}/reports/summary", timeout=30)
        assert r.status_code == 200
        assert r.json()["date"] == __import__("datetime").datetime.utcnow().strftime("%Y-%m-%d")

    def test_kasir_cannot_read_summary(self, kasir):
        assert kasir.get(f"{API}/reports/summary", timeout=30).status_code == 403

    def test_range_report(self, admin):
        import datetime
        today = datetime.date.today()
        r = admin.get(f"{API}/reports/range", params={"start": (today - datetime.timedelta(days=7)).isoformat(),
                                                      "end": today.isoformat()}, timeout=30)
        assert r.status_code == 200
        assert "daily" in r.json()

    def test_orders_list_filters(self, admin):
        import datetime
        r = admin.get(f"{API}/orders", params={"status": "paid",
                                               "date": datetime.date.today().isoformat()}, timeout=30)
        assert r.status_code == 200
        orders = r.json()
        assert all(o["status"] == "paid" for o in orders)
        for o in orders:
            assert "_id" not in o


# ---------------------------------------------------------------- EXCEL
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _build_xlsx(rows):
    import openpyxl
    cols = ["nama_produk", "sku", "kategori", "tipe_produk", "harga", "status_aktif",
            "sold_out", "deskripsi", "stok_awal"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(cols)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


class TestExcel:
    def test_template_download(self, admin):
        r = admin.get(f"{API}/products/template", timeout=60)
        assert r.status_code == 200, r.text[:200]
        assert XLSX_MIME in r.headers.get("content-type", "")
        assert r.content[:2] == b"PK"

    def test_export(self, admin):
        r = admin.get(f"{API}/products/export", timeout=60)
        assert r.status_code == 200, r.text[:200]
        assert r.content[:2] == b"PK"
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        rows = list(wb.active.iter_rows(values_only=True))
        assert rows[0][0] == "nama_produk"
        assert len(rows) > 1

    def test_import_preview_validations(self, admin, cat_map):
        retail_cat = cat_map["retail"]["name"]
        existing_sku = [p for p in admin.get(f"{API}/products", timeout=30).json()
                        if p["type"] == "retail"][0]["sku"]
        dup = _tag("SKUIMPDUP")
        rows = [
            [_tag("IMP_OK"), _tag("SKUIMP1"), retail_cat, "retail", 12000, "aktif", "tidak", "desc", 10],
            [_tag("IMP_DUPF1"), dup, retail_cat, "retail", 1000, "aktif", "tidak", "", 1],
            [_tag("IMP_DUPF2"), dup, retail_cat, "retail", 1000, "aktif", "tidak", "", 1],
            [_tag("IMP_BADCAT"), _tag("SKUIMP3"), "Kategori Palsu XYZ", "retail", 1000, "aktif", "tidak", "", 0],
            [_tag("IMP_BADTYPE"), _tag("SKUIMP4"), retail_cat, "elektronik", 1000, "aktif", "tidak", "", 0],
            [_tag("IMP_NEG"), _tag("SKUIMP5"), retail_cat, "retail", -500, "aktif", "tidak", "", 0],
            ["", _tag("SKUIMP6"), retail_cat, "retail", 1000, "aktif", "tidak", "", 0],
            [_tag("IMP_UPD"), existing_sku, retail_cat, "retail", 15500, "aktif", "tidak", "upd", 7],
        ]
        content = _build_xlsx(rows)
        r = admin.post(f"{API}/products/import/preview",
                       files={"file": ("t.xlsx", content, XLSX_MIME)}, timeout=60)
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["total"] == 8
        by_name = {row["name"]: row for row in d["rows"]}
        assert by_name[_tag("IMP_OK")]["valid"] is True
        assert any("duplikat" in e for e in by_name[_tag("IMP_DUPF2")]["errors"])
        assert any("kategori" in e for e in by_name[_tag("IMP_BADCAT")]["errors"])
        assert any("tipe" in e for e in by_name[_tag("IMP_BADTYPE")]["errors"])
        assert any("negatif" in e for e in by_name[_tag("IMP_NEG")]["errors"])
        empty_row = [row for row in d["rows"] if row["name"] == ""][0]
        assert any("nama" in e for e in empty_row["errors"])
        upd_row = by_name[_tag("IMP_UPD")]
        assert upd_row["exists"] is True and upd_row["valid"] is True
        assert d["error_count"] >= 5
        assert d["update_count"] >= 1
        assert d["new_count"] >= 1

    def test_import_commit_and_log(self, admin, cat_map):
        retail_cat = cat_map["retail"]["name"]
        sku_new = _tag("SKUCOMMIT")
        rows = [
            [_tag("IMP_COMMIT"), sku_new, retail_cat, "retail", 21000, "aktif", "tidak", "committed", 33],
            [_tag("IMP_BAD"), _tag("SKUCOMMITBAD"), "Kategori Palsu", "retail", 1000, "aktif", "tidak", "", 0],
        ]
        r = admin.post(f"{API}/products/import/commit",
                       files={"file": ("commit.xlsx", _build_xlsx(rows), XLSX_MIME)}, timeout=60)
        assert r.status_code == 200, r.text
        log = r.json()
        assert log["created"] == 1 and log["errors"] == 1
        created = [p for p in admin.get(f"{API}/products", timeout=30).json() if p["sku"] == sku_new]
        assert created, "committed product not found"
        assert created[0]["price"] == 21000 and created[0]["stock"] == 33
        assert created[0]["track_stock"] is True
        # update path
        rows2 = [[_tag("IMP_COMMIT_UPD"), sku_new, retail_cat, "retail", 22500, "aktif", "tidak", "upd", 40]]
        r2 = admin.post(f"{API}/products/import/commit",
                        files={"file": ("commit2.xlsx", _build_xlsx(rows2), XLSX_MIME)}, timeout=60)
        assert r2.status_code == 200, r2.text
        assert r2.json()["updated"] == 1
        upd = [p for p in admin.get(f"{API}/products", timeout=30).json() if p["sku"] == sku_new][0]
        assert upd["price"] == 22500 and upd["name"] == _tag("IMP_COMMIT_UPD")
        logs = admin.get(f"{API}/import-logs", timeout=30)
        assert logs.status_code == 200 and len(logs.json()) >= 1
        assert logs.json()[0]["filename"] in ("commit2.xlsx", "commit.xlsx")
        admin.delete(f"{API}/products/{upd['id']}", timeout=30)

    def test_kasir_cannot_import_or_export(self, kasir):
        assert kasir.get(f"{API}/products/export", timeout=30).status_code == 403
        r = kasir.post(f"{API}/products/import/preview",
                       files={"file": ("t.xlsx", _build_xlsx([]), XLSX_MIME)}, timeout=30)
        assert r.status_code == 403


# ---------------------------------------------------------------- AI
class TestAI:
    def test_ai_product_description(self, admin):
        r = admin.post(f"{API}/ai/product-description",
                       json={"name": "Mie Aceh Goreng Kepiting", "type": "makanan",
                             "category": "Makanan Utama", "keywords": "pedas, gurih"}, timeout=180)
        assert r.status_code == 200, r.text[:400]
        desc = r.json()["description"]
        assert isinstance(desc, str) and len(desc) > 20, desc

    def test_ai_product_image(self, admin):
        r = admin.post(f"{API}/ai/product-image",
                       json={"name": "Kopi Sanger", "description": "Kopi susu khas Aceh"}, timeout=240)
        assert r.status_code == 200, r.text[:400]
        img = r.json()["image"]
        assert img.startswith("data:image/") and ";base64," in img
        assert len(img) > 5000

    def test_ai_summary(self, admin):
        r = admin.post(f"{API}/reports/ai-summary", json={}, timeout=180)
        assert r.status_code == 200, r.text[:400]
        d = r.json()
        assert isinstance(d["summary"], str) and len(d["summary"]) > 30
        assert "data" in d

    def test_ai_requires_admin(self, kasir):
        r = kasir.post(f"{API}/ai/product-description",
                       json={"name": "x", "type": "makanan"}, timeout=60)
        assert r.status_code == 403


# ---------------------------------------------------------------- USERS
class TestUsers:
    def test_create_toggle_user(self, admin):
        email = f"test_{UQ}@example.com"
        r = admin.post(f"{API}/users", json={"name": _tag("USER"), "email": email,
                                             "password": "Passw0rd!23", "role": "kasir"}, timeout=30)
        assert r.status_code == 200, r.text
        uid = r.json()["id"]
        assert r.json()["role"] == "kasir"
        dup = admin.post(f"{API}/users", json={"name": "x", "email": email,
                                                "password": "Passw0rd!23"}, timeout=30)
        assert dup.status_code == 400
        # new user can login
        lg = requests.post(f"{API}/auth/login", json={"email": email, "password": "Passw0rd!23"}, timeout=30)
        assert lg.status_code == 200, lg.text
        t = admin.patch(f"{API}/users/{uid}/toggle", timeout=30)
        assert t.status_code == 200 and t.json()["active"] is False
        # deactivated user cannot login
        lg2 = requests.post(f"{API}/auth/login", json={"email": email, "password": "Passw0rd!23"}, timeout=30)
        assert lg2.status_code == 403, lg2.text
        admin.patch(f"{API}/users/{uid}/toggle", timeout=30)

    def test_toggle_nonexistent_user(self, admin):
        assert admin.patch(f"{API}/users/nope/toggle", timeout=30).status_code == 404


# ---------------------------------------------------------------- cleanup
# NOTE: cleanup is intentionally NOT a session-scoped autouse fixture. With pytest-xdist
# each worker ends its own session, so a global TEST_* purge would wipe data still in use
# by the other worker. Run `python /app/backend/tests/cleanup_test_data.py` after the suite.
