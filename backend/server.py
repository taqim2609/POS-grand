from dotenv import load_dotenv
from pathlib import Path
import os

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

UPLOAD_DIR = Path(os.environ.get('UPLOAD_DIR') or (ROOT_DIR / 'uploads'))
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
PROJECT_ROOT = ROOT_DIR.parent

from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, UploadFile, File, Query, BackgroundTasks
from fastapi.responses import StreamingResponse, FileResponse, Response
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pymongo import ReturnDocument
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional, Literal
from datetime import datetime, timezone, timedelta
import logging, uuid, io, bcrypt, jwt, asyncio, re, zipfile, math
from bson import json_util

# ------------------------------------------------------------------ DB
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

JWT_SECRET = os.environ['JWT_SECRET']
JWT_ALG = "HS256"
EMERGENT_LLM_KEY = os.environ.get('EMERGENT_LLM_KEY')
GEMINI_API_KEY = os.environ.get('GEMINI_API_KEY')
GEMINI_TEXT_MODEL = os.environ.get('GEMINI_MODEL', 'gemini-2.5-flash')
GEMINI_IMAGE_MODEL = os.environ.get('GEMINI_IMAGE_MODEL', 'gemini-2.5-flash-image')
OPENAI_COMPAT_BASE_URL = os.environ.get('OPENAI_COMPAT_BASE_URL')
OPENAI_COMPAT_API_KEY = os.environ.get('OPENAI_COMPAT_API_KEY')
OPENAI_COMPAT_MODEL = os.environ.get('OPENAI_COMPAT_MODEL') or 'gpt-4o-mini'
CHENZK_BASE_URL = "https://chenzk.top/v1"  # default base url provider "chenzk" (ezkielyna.store)
GEMINI_REST_URL = "https://generativelanguage.googleapis.com/v1beta/models"

app = FastAPI(title="Grand Aceh Kuliner POS")
api = APIRouter(prefix="/api")
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("gak-pos")

PRODUCT_TYPES = ["makanan", "minuman", "retail"]
ORDER_TYPES = ["dine_in", "take_away", "retail"]

def now_utc():
    return datetime.now(timezone.utc)

WIB = timezone(timedelta(hours=7))
LOW_STOCK_THRESHOLD = 10

def wib_today():
    return now_utc().astimezone(WIB).strftime("%Y-%m-%d")

def wib_day_range(date_str):
    """Given a WIB calendar date 'YYYY-MM-DD', return (start_utc_iso, end_utc_iso)."""
    try:
        start = datetime.strptime(date_str, "%Y-%m-%d").replace(tzinfo=WIB)
    except (ValueError, TypeError):
        raise HTTPException(400, f"Tanggal tidak valid: '{date_str}'. Gunakan format YYYY-MM-DD.")
    end = start + timedelta(days=1)
    return start.astimezone(timezone.utc).isoformat(), end.astimezone(timezone.utc).isoformat()

def wib_day_of(iso_str):
    return datetime.fromisoformat(iso_str).astimezone(WIB).strftime("%Y-%m-%d")

def new_id():
    return str(uuid.uuid4())

# ------------------------------------------------------------------ Security
def hash_password(p: str) -> str:
    return bcrypt.hashpw(p.encode(), bcrypt.gensalt()).decode()

def verify_password(p: str, h: str) -> bool:
    try:
        return bcrypt.checkpw(p.encode(), h.encode())
    except Exception:
        return False

def create_token(user: dict) -> str:
    payload = {"sub": user["id"], "role": user["role"], "email": user["email"],
               "exp": now_utc() + timedelta(hours=12), "type": "access"}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALG)

async def get_current_user(request: Request) -> dict:
    auth = request.headers.get("Authorization", "")
    token = auth[7:] if auth.startswith("Bearer ") else request.cookies.get("access_token")
    if not token:
        raise HTTPException(401, "Not authenticated")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALG])
    except jwt.ExpiredSignatureError:
        raise HTTPException(401, "Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(401, "Invalid token")
    user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
    if not user or not user.get("active", True):
        raise HTTPException(401, "User not found or inactive")
    return user

async def require_admin(user: dict = Depends(get_current_user)) -> dict:
    if user["role"] != "admin":
        raise HTTPException(403, "Admin access required")
    return user

def require_roles(*roles):
    async def _dep(user: dict = Depends(get_current_user)) -> dict:
        if user["role"] not in roles:
            raise HTTPException(403, "Akses ditolak untuk peran ini")
        return user
    return _dep

admin_or_input = require_roles("admin", "input")
admin_or_kasir = require_roles("admin", "kasir")

# ------------------------------------------------------------------ Models
class LoginIn(BaseModel):
    email: EmailStr
    password: str

class UserCreate(BaseModel):
    name: str
    email: EmailStr
    password: str
    role: Literal["admin", "kasir", "input"] = "kasir"

class ChangePasswordIn(BaseModel):
    current_password: str
    new_password: str

class ResetPasswordIn(BaseModel):
    new_password: str

class ResetDataIn(BaseModel):
    scope: Literal["transactions", "all"]
    password: str

class CategoryIn(BaseModel):
    name: str
    type: Literal["makanan", "minuman", "retail", "vendor"]
    sort_order: int = 0
    active: bool = True

class ProductIn(BaseModel):
    name: str
    sku: str
    category_id: str
    type: Literal["makanan", "minuman", "retail", "vendor"]
    price: float
    cost: float = 0
    vendor_id: Optional[str] = None
    vendor_share_percent: Optional[float] = None
    description: Optional[str] = ""
    image: Optional[str] = ""
    active: bool = True
    sold_out: bool = False
    stock: Optional[int] = 0
    min_stock: Optional[int] = 10

class VendorIn(BaseModel):
    name: str
    contact: Optional[str] = ""
    note: Optional[str] = ""
    active: bool = True

class TableIn(BaseModel):
    name: str
    area: str = "Umum"
    capacity: int = 4
    active: bool = True

class PaymentMethodIn(BaseModel):
    name: str
    type: Literal["cash", "qris", "card"]
    active: bool = True

class OrderItem(BaseModel):
    product_id: str
    name: Optional[str] = ""
    price: Optional[float] = 0
    qty: int = Field(gt=0)
    type: Optional[str] = ""

class OrderIn(BaseModel):
    order_type: Literal["dine_in", "take_away", "retail"]
    table_id: Optional[str] = None
    items: List[OrderItem]
    discount_type: Literal["none", "percent", "amount"] = "none"
    discount_value: float = Field(0, ge=0)
    note: Optional[str] = ""
    pay_now: bool = False
    payment_method: Optional[str] = None
    client_ref: Optional[str] = None
    member_id: Optional[str] = None
    redeem_points: float = 0
    discount_reason: Optional[str] = None

class ItemsUpdate(BaseModel):
    items: List[OrderItem]

class PayIn(BaseModel):
    payment_method: str
    discount_type: Literal["none", "percent", "amount"] = "none"
    discount_value: float = Field(0, ge=0)
    amount_paid: Optional[float] = None
    member_id: Optional[str] = None
    redeem_points: float = 0
    discount_reason: Optional[str] = None

class VoidIn(BaseModel):
    reason: str
    action: Literal["void", "refund"] = "void"

class ShiftOpenIn(BaseModel):
    opening_cash: float = 0

class ShiftCloseIn(BaseModel):
    closing_cash: float = 0
    vendor_payments: Optional[List[dict]] = None  # [{vendor_id, paid}]

class AIDescIn(BaseModel):
    name: str
    type: str
    category: Optional[str] = ""
    keywords: Optional[str] = ""

class AIImageIn(BaseModel):
    name: str
    description: Optional[str] = ""

class AISummaryIn(BaseModel):
    date: Optional[str] = None

# ------------------------------------------------------------------ helpers
def compute_totals(items, discount_type, discount_value):
    subtotal = sum(i["price"] * i["qty"] for i in items)
    discount_value = max(0, discount_value)
    if discount_type == "percent":
        discount = round(subtotal * (min(discount_value, 100) / 100.0), 2)
    elif discount_type == "amount":
        discount = min(discount_value, subtotal)
    else:
        discount = 0
    total = max(0, round(subtotal - discount, 2))
    return round(subtotal, 2), round(discount, 2), total

# ================================================================== PROMO & MEMBER (fitur baru)
async def _active_promos():
    return await db.promos.find({"active": True}, {"_id": 0}).to_list(100)

async def _apply_promos(items, subtotal, now=None):
    """Hitung diskon otomatis dari promo aktif. Return (promo_discount, [nama_promo])."""
    now = now or datetime.now(WIB)
    t = now.strftime("%H:%M")
    wd = now.weekday()  # 0=Senin..6=Minggu
    names = []
    d = 0.0
    for p in await _active_promos():
        try:
            days = p.get("days") or []
            if days and wd not in days:
                continue
            ptype = p.get("type")
            if ptype in ("percent", "happy_hour"):
                if ptype == "happy_hour":
                    st, en = p.get("start_time", ""), p.get("end_time", "")
                    if not (st and en) or not (st <= t <= en):
                        continue
                pct = min(float(p.get("value") or 0), 100)
                d += round(subtotal * pct / 100.0, 2)
                names.append(p.get("name"))
            elif ptype == "min_spend":
                if subtotal >= float(p.get("value") or 0):
                    bonus = float(p.get("bonus") or 0)
                    d += min(bonus, subtotal)
                    names.append(p.get("name"))
            elif ptype == "package":
                # paket: semua produk dalam paket harus ada di keranjang (cocokkan nama)
                pkg = p.get("package_items") or []
                ok = True
                for need in pkg:
                    nm = (need.get("product_name") or "").lower().strip()
                    q = int(need.get("qty") or 1)
                    if nm:
                        have = sum(i["qty"] for i in items if (i["name"] or "").lower().strip() == nm)
                        if have < q:
                            ok = False
                            break
                if ok:
                    bundle = float(p.get("value") or 0)
                    normal = 0.0
                    for need in pkg:
                        nm = (need.get("product_name") or "").lower().strip()
                        q = int(need.get("qty") or 1)
                        for i in items:
                            if (i["name"] or "").lower().strip() == nm:
                                normal += i["price"] * q
                                break
                    d += max(0.0, normal - bundle)
                    names.append(p.get("name"))
            elif ptype == "bogo":
                # beli N gratis 1 (produk sama, unit paling murah dihitung)
                buy = int(p.get("value") or 2)
                for it in items:
                    if it["qty"] >= buy + 1:
                        free = it["price"]
                        d += min(free, subtotal)
                        names.append(p.get("name"))
                        break
        except Exception:
            continue
    promo_discount = min(d, subtotal)
    return round(promo_discount, 2), names

def _discount_needs_reason(discount_type, discount_value, subtotal):
    if discount_type == "percent":
        return discount_value > 15
    if discount_type == "amount":
        return discount_value > 50000
    return False

async def _apply_redeem(member_id, redeem_points, total_before):
    """Validasi & potong poin member; return potongan rupiah (1 poin = Rp100)."""
    if not member_id or not redeem_points or redeem_points <= 0:
        return 0.0, None
    m = await db.members.find_one({"id": member_id})
    if not m:
        raise HTTPException(400, "Member tidak ditemukan")
    pts = float(m.get("points") or 0)
    if redeem_points > pts:
        raise HTTPException(400, f"Poin member tidak cukup (sisa {pts:,.0f})")
    value = round(redeem_points * 100, 2)
    value = min(value, total_before)
    await db.members.update_one({"id": member_id}, {"$inc": {"points": -redeem_points}})
    return value, m

async def _award_member_points(order):
    """Setelah order lunas, tambah poin & total belanja member + kirim notifikasi WA."""
    mid = order.get("member_id")
    if not mid:
        return
    m = await db.members.find_one({"id": mid})
    if not m:
        return
    pts = int(order.get("total", 0) // 10000)  # 1 poin per Rp10.000
    if pts <= 0:
        return
    await db.members.update_one({"id": mid}, {"$inc": {"points": pts, "total_spend": order.get("total", 0)}})
    await db.orders.update_one({"id": order["id"]}, {"$set": {"points_earned": pts}})
    phone = m.get("phone", "")
    if phone and await _wa_configured():
        try:
            await _send_whatsapp([phone],
                f"*Grand Aceh Kuliner*\nTerima kasih sudah berbelanja!\nTotal: Rp{order.get('total',0):,.0f}\nPoin +{pts} (total poin {float(m.get('points') or 0)+pts:,.0f})\nTukarkan poin 1pt=Rp100 saat pembayaran.")
        except Exception:
            pass

async def gen_order_number():
    today = wib_today().replace("-", "")
    cid = f"order-{today}"
    # migration-safe init: seed counter from any orders already created today
    if not await db.counters.find_one({"_id": cid}):
        cnt = await db.orders.count_documents({"order_number": {"$regex": f"^GAK-{today}-"}})
        await db.counters.update_one({"_id": cid}, {"$setOnInsert": {"seq": cnt}}, upsert=True)
    doc = await db.counters.find_one_and_update(
        {"_id": cid}, {"$inc": {"seq": 1}}, return_document=ReturnDocument.AFTER
    )
    return f"GAK-{today}-{doc['seq']:04d}"

# ================================================================== AUTH
@api.post("/auth/login")
async def login(body: LoginIn):
    email = body.email.lower()
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(body.password, user["password_hash"]):
        raise HTTPException(401, "Email atau password salah")
    if not user.get("active", True):
        raise HTTPException(403, "Akun dinonaktifkan")
    safe = {"id": user["id"], "name": user["name"], "email": user["email"], "role": user["role"]}
    return {"token": create_token(safe), "user": safe}

@api.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    return user

@api.get("/users")
async def list_users(admin: dict = Depends(require_admin)):
    return await db.users.find({}, {"_id": 0, "password_hash": 0}).sort("created_at", 1).to_list(500)

@api.post("/users")
async def create_user(body: UserCreate, admin: dict = Depends(require_admin)):
    if await db.users.find_one({"email": body.email.lower()}):
        raise HTTPException(400, "Email sudah dipakai")
    doc = {"id": new_id(), "name": body.name, "email": body.email.lower(),
           "password_hash": hash_password(body.password), "role": body.role,
           "active": True, "created_at": now_utc().isoformat()}
    await db.users.insert_one(doc)
    return {"id": doc["id"], "name": doc["name"], "email": doc["email"], "role": doc["role"]}

@api.patch("/users/{uid}/toggle")
async def toggle_user(uid: str, admin: dict = Depends(require_admin)):
    u = await db.users.find_one({"id": uid})
    if not u:
        raise HTTPException(404, "User tidak ditemukan")
    await db.users.update_one({"id": uid}, {"$set": {"active": not u.get("active", True)}})
    return {"active": not u.get("active", True)}

@api.post("/auth/change-password")
async def change_password(body: ChangePasswordIn, user: dict = Depends(get_current_user)):
    if len(body.new_password) < 6:
        raise HTTPException(400, "Password baru minimal 6 karakter")
    full = await db.users.find_one({"id": user["id"]})
    if not full or not verify_password(body.current_password, full["password_hash"]):
        raise HTTPException(400, "Password lama salah")
    await db.users.update_one({"id": user["id"]}, {"$set": {"password_hash": hash_password(body.new_password)}})
    return {"ok": True}

@api.post("/users/{uid}/reset-password")
async def reset_user_password(uid: str, body: ResetPasswordIn, admin: dict = Depends(require_admin)):
    if len(body.new_password) < 6:
        raise HTTPException(400, "Password minimal 6 karakter")
    u = await db.users.find_one({"id": uid})
    if not u:
        raise HTTPException(404, "User tidak ditemukan")
    await db.users.update_one({"id": uid}, {"$set": {"password_hash": hash_password(body.new_password)}})
    return {"ok": True}

@api.post("/admin/reset-data")
async def reset_data(body: ResetDataIn, admin: dict = Depends(require_admin)):
    """Destructive: wipe transactional (and optionally catalog) data. Keeps users, settings, payment methods."""
    full = await db.users.find_one({"id": admin["id"]})
    if not full or not verify_password(body.password, full["password_hash"]):
        raise HTTPException(400, "Password admin salah")
    tx = ["orders", "cash_movements", "shifts", "stock_opname", "purchases", "counters", "import_logs", "audit_logs"]
    catalog = ["products", "categories", "tables"]
    cols = tx + (catalog if body.scope == "all" else [])
    deleted = {}
    for col in cols:
        r = await db[col].delete_many({})
        deleted[col] = r.deleted_count
    return {"ok": True, "scope": body.scope, "deleted": deleted}

# ================================================================== CATEGORIES
@api.get("/categories")
async def list_categories(include_inactive: bool = True, user: dict = Depends(get_current_user)):
    q = {} if include_inactive else {"active": True}
    return await db.categories.find(q, {"_id": 0}).sort("sort_order", 1).to_list(500)

@api.post("/categories")
async def create_category(body: CategoryIn, admin: dict = Depends(admin_or_input)):
    if await db.categories.find_one({"name": {"$regex": f"^{re.escape(body.name.strip())}$", "$options": "i"}, "type": body.type}):
        raise HTTPException(400, f"Kategori '{body.name}' sudah ada untuk tipe ini")
    doc = body.model_dump()
    doc.update({"id": new_id(), "created_at": now_utc().isoformat()})
    await db.categories.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.put("/categories/{cid}")
async def update_category(cid: str, body: CategoryIn, admin: dict = Depends(admin_or_input)):
    if not await db.categories.find_one({"id": cid}):
        raise HTTPException(404, "Kategori tidak ditemukan")
    if await db.categories.find_one({"name": {"$regex": f"^{re.escape(body.name.strip())}$", "$options": "i"}, "type": body.type, "id": {"$ne": cid}}):
        raise HTTPException(400, f"Kategori '{body.name}' sudah ada untuk tipe ini")
    await db.categories.update_one({"id": cid}, {"$set": body.model_dump()})
    return await db.categories.find_one({"id": cid}, {"_id": 0})

@api.delete("/categories/{cid}")
async def delete_category(cid: str, admin: dict = Depends(admin_or_input)):
    used = await db.products.count_documents({"category_id": cid})
    if used:
        # soft deactivate instead of hard delete
        await db.categories.update_one({"id": cid}, {"$set": {"active": False}})
        return {"soft_deleted": True, "reason": f"Kategori dipakai {used} produk, dinonaktifkan (tidak dihapus)."}
    await db.categories.delete_one({"id": cid})
    return {"deleted": True}

# ================================================================== PRODUCTS
@api.get("/products")
async def list_products(type: Optional[str] = None, category_id: Optional[str] = None,
                        active_only: bool = False, user: dict = Depends(get_current_user)):
    q = {}
    if type:
        q["type"] = type
    if category_id:
        q["category_id"] = category_id
    if active_only:
        q["active"] = True
    return await db.products.find(q, {"_id": 0}).sort("name", 1).to_list(2000)

@api.post("/products")
async def create_product(body: ProductIn, admin: dict = Depends(admin_or_input)):
    if body.price < 0 or body.cost < 0:
        raise HTTPException(400, "Harga/HPP tidak boleh negatif")
    if await db.products.find_one({"sku": body.sku}):
        raise HTTPException(400, f"SKU '{body.sku}' sudah dipakai")
    if not await db.categories.find_one({"id": body.category_id}):
        raise HTTPException(400, "Kategori tidak valid")
    doc = body.model_dump()
    doc["track_stock"] = body.type == "retail"
    doc.update({"id": new_id(), "created_at": now_utc().isoformat()})
    await db.products.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.put("/products/{pid}")
async def update_product(pid: str, body: ProductIn, admin: dict = Depends(admin_or_input)):
    if body.price < 0 or body.cost < 0:
        raise HTTPException(400, "Harga/HPP tidak boleh negatif")
    existing = await db.products.find_one({"id": pid})
    if not existing:
        raise HTTPException(404, "Produk tidak ditemukan")
    dup = await db.products.find_one({"sku": body.sku, "id": {"$ne": pid}})
    if dup:
        raise HTTPException(400, f"SKU '{body.sku}' sudah dipakai produk lain")
    doc = body.model_dump(exclude_unset=True)
    if "type" in doc:
        doc["track_stock"] = doc["type"] == "retail"
    await db.products.update_one({"id": pid}, {"$set": doc})
    return await db.products.find_one({"id": pid}, {"_id": 0})

@api.patch("/products/{pid}/sold-out")
async def toggle_sold_out(pid: str, user: dict = Depends(get_current_user)):
    p = await db.products.find_one({"id": pid})
    if not p:
        raise HTTPException(404, "Produk tidak ditemukan")
    val = not p.get("sold_out", False)
    await db.products.update_one({"id": pid}, {"$set": {"sold_out": val}})
    return {"sold_out": val}

@api.delete("/products/{pid}")
async def delete_product(pid: str, admin: dict = Depends(admin_or_input)):
    used = await db.orders.count_documents({"items.product_id": pid})
    if used:
        await db.products.update_one({"id": pid}, {"$set": {"active": False}})
        return {"soft_deleted": True, "reason": "Produk pernah dipakai transaksi, dinonaktifkan."}
    await db.products.delete_one({"id": pid})
    return {"deleted": True}

# ================================================================== VENDORS
@api.get("/vendors")
async def list_vendors(active_only: bool = False, user: dict = Depends(get_current_user)):
    q = {"active": True} if active_only else {}
    return await db.vendors.find(q, {"_id": 0}).sort("name", 1).to_list(500)

@api.post("/vendors")
async def create_vendor(body: VendorIn, admin: dict = Depends(admin_or_input)):
    if await db.vendors.find_one({"name": {"$regex": f"^{re.escape(body.name.strip())}$", "$options": "i"}}):
        raise HTTPException(400, f"Vendor '{body.name}' sudah ada")
    doc = body.model_dump()
    doc.update({"id": new_id(), "created_at": now_utc().isoformat()})
    await db.vendors.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.put("/vendors/{vid}")
async def update_vendor(vid: str, body: VendorIn, admin: dict = Depends(admin_or_input)):
    if not await db.vendors.find_one({"id": vid}):
        raise HTTPException(404, "Vendor tidak ditemukan")
    if await db.vendors.find_one({"name": {"$regex": f"^{re.escape(body.name.strip())}$", "$options": "i"}, "id": {"$ne": vid}}):
        raise HTTPException(400, f"Vendor '{body.name}' sudah ada")
    await db.vendors.update_one({"id": vid}, {"$set": body.model_dump()})
    return await db.vendors.find_one({"id": vid}, {"_id": 0})

@api.delete("/vendors/{vid}")
async def delete_vendor(vid: str, admin: dict = Depends(admin_or_input)):
    used = await db.products.count_documents({"vendor_id": vid})
    if used:
        await db.vendors.update_one({"id": vid}, {"$set": {"active": False}})
        return {"soft_deleted": True, "reason": f"Vendor dipakai {used} produk, dinonaktifkan (tidak dihapus)."}
    await db.vendors.delete_one({"id": vid})
    return {"deleted": True}

# ================================================================== TABLES
@api.get("/tables")
async def list_tables(user: dict = Depends(get_current_user)):
    tables = await db.tables.find({"deleted": {"$ne": True}}, {"_id": 0}).sort("area", 1).to_list(500)
    open_orders = await db.orders.find({"order_type": "dine_in", "status": "open"}, {"_id": 0, "table_id": 1, "id": 1, "total": 1}).to_list(1000)
    open_map = {}
    for o in open_orders:
        open_map[o["table_id"]] = o
    for t in tables:
        oo = open_map.get(t["id"])
        t["status"] = "open_bill" if oo else "empty"
        t["open_order_id"] = oo["id"] if oo else None
    return tables

@api.post("/tables")
async def create_table(body: TableIn, admin: dict = Depends(require_admin)):
    if await db.tables.find_one({"name": body.name, "deleted": {"$ne": True}}):
        raise HTTPException(400, f"Nama/kode meja '{body.name}' sudah ada")
    doc = body.model_dump()
    doc.update({"id": new_id(), "deleted": False, "created_at": now_utc().isoformat()})
    await db.tables.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.put("/tables/{tid}")
async def update_table(tid: str, body: TableIn, admin: dict = Depends(require_admin)):
    t = await db.tables.find_one({"id": tid})
    if not t:
        raise HTTPException(404, "Meja tidak ditemukan")
    dup = await db.tables.find_one({"name": body.name, "id": {"$ne": tid}, "deleted": {"$ne": True}})
    if dup:
        raise HTTPException(400, f"Nama/kode meja '{body.name}' sudah ada")
    if t.get("active") and not body.active:
        open_bill = await db.orders.find_one({"table_id": tid, "status": "open"})
        if open_bill:
            raise HTTPException(400, "Meja punya open bill aktif, tidak bisa dinonaktifkan")
    await db.tables.update_one({"id": tid}, {"$set": body.model_dump()})
    return await db.tables.find_one({"id": tid}, {"_id": 0})

@api.delete("/tables/{tid}")
async def delete_table(tid: str, admin: dict = Depends(require_admin)):
    open_bill = await db.orders.find_one({"table_id": tid, "status": "open"})
    if open_bill:
        raise HTTPException(400, "Meja punya open bill aktif, tidak bisa dihapus")
    used = await db.orders.count_documents({"table_id": tid})
    if used:
        await db.tables.update_one({"id": tid}, {"$set": {"active": False}})
        return {"soft_deleted": True, "reason": "Meja pernah dipakai transaksi, dinonaktifkan (tidak dihapus)."}
    await db.tables.update_one({"id": tid}, {"$set": {"deleted": True, "active": False}})
    return {"deleted": True}

# ================================================================== PAYMENT METHODS
@api.get("/payment-methods")
async def list_payment_methods(user: dict = Depends(get_current_user)):
    return await db.payment_methods.find({}, {"_id": 0}).sort("name", 1).to_list(100)

@api.post("/payment-methods")
async def create_pm(body: PaymentMethodIn, admin: dict = Depends(require_admin)):
    doc = body.model_dump()
    doc.update({"id": new_id()})
    await db.payment_methods.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.patch("/payment-methods/{pmid}/toggle")
async def toggle_pm(pmid: str, admin: dict = Depends(require_admin)):
    pm = await db.payment_methods.find_one({"id": pmid})
    if not pm:
        raise HTTPException(404, "Metode tidak ditemukan")
    await db.payment_methods.update_one({"id": pmid}, {"$set": {"active": not pm.get("active", True)}})
    return {"active": not pm.get("active", True)}

# ================================================================== ORDERS
async def _resolve_items(raw_items):
    """Rebuild every line item from the products collection (server-authoritative)."""
    resolved = []
    for it in raw_items:
        p = await db.products.find_one({"id": it.product_id}, {"_id": 0})
        if not p:
            raise HTTPException(400, f"Produk tidak ditemukan: {it.product_id}")
        if not p.get("active", True):
            raise HTTPException(400, f"Produk nonaktif tidak bisa dijual: {p['name']}")
        if p.get("sold_out"):
            raise HTTPException(400, f"Produk sold out: {p['name']}")
        line = {"product_id": p["id"], "name": p["name"], "price": p["price"],
                "cost": p.get("cost", 0),
                "qty": it.qty, "type": p["type"], "track_stock": p.get("track_stock", False)}
        if p["type"] == "vendor" and p.get("vendor_id"):
            share = float(p.get("vendor_share_percent") or 0)
            line["vendor_id"] = p["vendor_id"]
            line["vendor_share_percent"] = share
            line["vendor_total"] = round(p["price"] * it.qty * share / 100, 2)
        resolved.append(line)
    return resolved

async def _validate_order_rules(order_type, table_id, items):
    types = {i["type"] for i in items}
    if order_type == "retail":
        if any(t != "retail" for t in types):
            raise HTTPException(400, "Alur retail hanya boleh berisi item retail")
    else:  # dine_in / take_away
        if "retail" in types:
            raise HTTPException(400, "Item retail tidak boleh masuk alur F&B (dine-in/take away)")
    if order_type == "dine_in":
        if not table_id:
            raise HTTPException(400, "Dine-in wajib memilih meja")
        t = await db.tables.find_one({"id": table_id, "deleted": {"$ne": True}})
        if not t or not t.get("active", True):
            raise HTTPException(400, "Meja tidak valid atau nonaktif")
    elif table_id:
        raise HTTPException(400, "Take away/retail tidak boleh memakai meja")

async def _current_shift(user):
    return await db.shifts.find_one({"cashier_id": user["id"], "status": "open"}, {"_id": 0})

async def _finalize_payment(order, payment_method, amount_paid, user):
    pm = await db.payment_methods.find_one({"id": payment_method, "active": True})
    if not pm:
        raise HTTPException(400, "Metode pembayaran tidak valid/aktif")
    paid = amount_paid if amount_paid is not None else order["total"]
    if pm["type"] == "cash" and paid < order["total"]:
        raise HTTPException(400, "Jumlah bayar kurang dari total")
    shift = await _current_shift(user)
    # validate & decrement retail stock atomically
    for it in order["items"]:
        if it.get("type") == "retail":
            p = await db.products.find_one({"id": it["product_id"]}, {"_id": 0})
            if p and p.get("track_stock") and p.get("stock", 0) < it["qty"]:
                raise HTTPException(400, f"Stok '{it['name']}' tidak cukup (sisa {p.get('stock', 0)})")
    for it in order["items"]:
        if it.get("type") == "retail":
            await db.products.update_one({"id": it["product_id"], "track_stock": True},
                                         {"$inc": {"stock": -it["qty"]}})
    upd = {"status": "paid", "payment_method_id": payment_method, "payment_method_name": pm["name"],
           "payment_method_type": pm["type"], "amount_paid": paid,
           "change": round(paid - order["total"], 2),
           "paid_at": now_utc().isoformat(), "shift_id": shift["id"] if shift else None}
    await db.orders.update_one({"id": order["id"]}, {"$set": upd})
    await _award_member_points({**order, **upd})
    return {**order, **upd}

@api.post("/orders")
async def create_order(body: OrderIn, user: dict = Depends(get_current_user)):
    if not body.items:
        raise HTTPException(400, "Keranjang kosong")
    if body.client_ref:
        dup = await db.orders.find_one({"client_ref": body.client_ref}, {"_id": 0})
        if dup:
            return dup  # idempotent: offline sync retry won't duplicate
    items = await _resolve_items(body.items)
    await _validate_order_rules(body.order_type, body.table_id, items)
    subtotal, discount, total = compute_totals(items, body.discount_type, body.discount_value)
    promo_discount, promo_names = await _apply_promos(items, subtotal)
    if _discount_needs_reason(body.discount_type, body.discount_value, subtotal) and not (body.discount_reason or "").strip():
        raise HTTPException(400, "Alasan diskon wajib diisi untuk diskon besar (>15% atau >Rp50.000)")
    total = max(0.0, round(total - promo_discount, 2))
    redeem_discount = 0.0
    if body.member_id and body.redeem_points:
        rd, m = await _apply_redeem(body.member_id, body.redeem_points, total)
        redeem_discount = rd
        total = max(0.0, round(total - rd, 2))
    doc = {
        "id": new_id(), "order_number": await gen_order_number(),
        "order_type": body.order_type, "table_id": body.table_id, "items": items,
        "subtotal": subtotal, "discount_type": body.discount_type, "discount_value": body.discount_value,
        "discount": discount, "promo_discount": promo_discount, "promos_applied": promo_names,
        "redeem_discount": redeem_discount, "member_id": body.member_id,
        "discount_reason": body.discount_reason, "total": total, "note": body.note,
        "status": "open", "cashier_id": user["id"], "cashier_name": user["name"],
        "client_ref": body.client_ref,
        "created_at": now_utc().isoformat(),
    }
    await db.orders.insert_one(doc)
    doc.pop("_id", None)
    if body.pay_now:
        if not body.payment_method:
            raise HTTPException(400, "Pilih metode pembayaran")
        doc = await _finalize_payment(doc, body.payment_method, None, user)
    return doc

@api.get("/orders")
async def list_orders(status: Optional[str] = None, order_type: Optional[str] = None,
                      date_str: Optional[str] = Query(None, alias="date"),
                      user: dict = Depends(get_current_user)):
    q = {}
    if status:
        q["status"] = status
    if order_type:
        q["order_type"] = order_type
    if date_str:
        start, end = wib_day_range(date_str)
        q["created_at"] = {"$gte": start, "$lt": end}
    return await db.orders.find(q, {"_id": 0}).sort("created_at", -1).to_list(1000)

@api.get("/orders/{oid}")
async def get_order(oid: str, user: dict = Depends(get_current_user)):
    o = await db.orders.find_one({"id": oid}, {"_id": 0})
    if not o:
        raise HTTPException(404, "Order tidak ditemukan")
    return o

@api.patch("/orders/{oid}/items")
async def update_order_items(oid: str, body: ItemsUpdate, user: dict = Depends(get_current_user)):
    o = await db.orders.find_one({"id": oid})
    if not o:
        raise HTTPException(404, "Order tidak ditemukan")
    if o["status"] != "open":
        raise HTTPException(400, "Hanya open bill yang bisa ditambah item")
    items = await _resolve_items(body.items)
    await _validate_order_rules(o["order_type"], o.get("table_id"), items)
    subtotal, discount, total = compute_totals(items, o["discount_type"], o["discount_value"])
    promo_discount, promo_names = await _apply_promos(items, subtotal)
    total = max(0.0, round(total - promo_discount, 2))
    await db.orders.update_one({"id": oid}, {"$set": {"items": items, "subtotal": subtotal,
                                                       "discount": discount, "promo_discount": promo_discount,
                                                       "promos_applied": promo_names, "total": total}})
    return await db.orders.find_one({"id": oid}, {"_id": 0})

@api.post("/orders/{oid}/pay")
async def pay_order(oid: str, body: PayIn, user: dict = Depends(get_current_user)):
    o = await db.orders.find_one({"id": oid}, {"_id": 0})
    if not o:
        raise HTTPException(404, "Order tidak ditemukan")
    if o["status"] != "open":
        raise HTTPException(400, "Order sudah lunas / tidak bisa dibayar")
    subtotal, discount, total = compute_totals(o["items"], body.discount_type, body.discount_value)
    promo_discount, promo_names = await _apply_promos(o["items"], subtotal)
    if _discount_needs_reason(body.discount_type, body.discount_value, subtotal) and not (body.discount_reason or "").strip():
        raise HTTPException(400, "Alasan diskon wajib diisi untuk diskon besar (>15% atau >Rp50.000)")
    total = max(0.0, round(total - promo_discount, 2))
    redeem_discount = 0.0
    if body.member_id and body.redeem_points:
        rd, m = await _apply_redeem(body.member_id, body.redeem_points, total)
        redeem_discount = rd
        total = max(0.0, round(total - rd, 2))
    await db.orders.update_one({"id": oid}, {"$set": {"discount_type": body.discount_type,
                                                      "discount_value": body.discount_value,
                                                      "discount": discount, "promo_discount": promo_discount,
                                                      "promos_applied": promo_names,
                                                      "redeem_discount": redeem_discount,
                                                      "member_id": body.member_id,
                                                      "discount_reason": body.discount_reason,
                                                      "total": total, "subtotal": subtotal}})
    o.update({"discount": discount, "promo_discount": promo_discount, "promos_applied": promo_names,
              "redeem_discount": redeem_discount, "member_id": body.member_id,
              "discount_reason": body.discount_reason, "total": total, "subtotal": subtotal})
    return await _finalize_payment(o, body.payment_method, body.amount_paid, user)

@api.post("/orders/{oid}/void")
async def void_order(oid: str, body: VoidIn, admin: dict = Depends(require_admin)):
    o = await db.orders.find_one({"id": oid})
    if not o:
        raise HTTPException(404, "Order tidak ditemukan")
    if o["status"] in ("void", "refunded"):
        raise HTTPException(400, "Order sudah dibatalkan/refund")
    new_status = "refunded" if body.action == "refund" else "void"
    # restore retail stock if it was paid
    if o["status"] == "paid":
        for it in o["items"]:
            if it["type"] == "retail":
                await db.products.update_one({"id": it["product_id"], "track_stock": True},
                                             {"$inc": {"stock": it["qty"]}})
    audit = {"id": new_id(), "order_id": oid, "order_number": o["order_number"],
             "action": body.action, "reason": body.reason, "prev_status": o["status"],
             "by": admin["name"], "by_id": admin["id"], "amount": o["total"],
             "at": now_utc().isoformat()}
    await db.audit_logs.insert_one(audit)
    await db.orders.update_one({"id": oid}, {"$set": {"status": new_status,
                                                      "voided_at": now_utc().isoformat(),
                                                      "void_reason": body.reason, "voided_by": admin["name"]}})
    audit.pop("_id", None)
    return {"status": new_status, "audit": audit}

@api.get("/audit-logs")
async def audit_logs(admin: dict = Depends(require_admin)):
    return await db.audit_logs.find({}, {"_id": 0}).sort("at", -1).to_list(500)

# ================================================================== SHIFTS
@api.get("/shifts/current")
async def current_shift(user: dict = Depends(get_current_user)):
    return await _current_shift(user)

@api.post("/shifts/open")
async def open_shift(body: ShiftOpenIn, user: dict = Depends(get_current_user)):
    if await _current_shift(user):
        raise HTTPException(400, "Sudah ada shift terbuka")
    doc = {"id": new_id(), "cashier_id": user["id"], "cashier_name": user["name"],
           "opening_cash": body.opening_cash, "status": "open",
           "opened_at": now_utc().isoformat(), "closed_at": None}
    await db.shifts.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.post("/shifts/close")
async def close_shift(body: ShiftCloseIn, user: dict = Depends(get_current_user)):
    shift = await _current_shift(user)
    if not shift:
        raise HTTPException(400, "Tidak ada shift terbuka")
    report = await _shift_report(shift)
    # Hitung bagian vendor (yang seharusnya) & yang diberikan dari input
    vendor_rows = report.get("vendor_share", [])
    paid_map = {str(v.get("vendor_id")): float(v.get("paid") or 0) for v in (body.vendor_payments or [])}
    total_share = 0.0
    total_paid = 0.0
    for v in vendor_rows:
        v["paid"] = paid_map.get(str(v["vendor_id"]), 0.0)
        v["difference"] = round(v["share"] - v["paid"], 2)
        total_share += v["share"]
        total_paid += v["paid"]
    total_diff = round(total_share - total_paid, 2)
    net_cash = round(report["expected_cash"] - total_paid, 2)  # uang bersih setelah bayar bagian vendor
    report["vendor_share"] = vendor_rows
    report["vendor_total_share"] = round(total_share, 2)
    report["vendor_total_paid"] = round(total_paid, 2)
    report["vendor_total_difference"] = total_diff
    report["net_cash"] = net_cash
    await db.shifts.update_one({"id": shift["id"]}, {"$set": {
        "status": "closed", "closed_at": now_utc().isoformat(),
        "closing_cash": body.closing_cash, "report": report}})
    return {**shift, "closing_cash": body.closing_cash, "report": report, "status": "closed"}

async def _vendor_share_from_orders(orders):
    """Kumpulkan bagian vendor dari order (per vendor): gross & share (yang seharusnya)."""
    per = {}
    for o in orders:
        for it in o.get("items", []):
            if it.get("type") == "vendor" and it.get("vendor_id"):
                vid = it["vendor_id"]
                v = per.get(vid)
                if not v:
                    v = {"vendor_id": vid, "vendor_name": it.get("vendor_name") or "Vendor", "gross": 0.0, "share": 0.0}
                    per[vid] = v
                v["gross"] = round(v["gross"] + it["price"] * it["qty"], 2)
                v["share"] = round(v["share"] + (it.get("vendor_total") or 0), 2)
    # isi nama vendor dari koleksi bila belum ada di item
    for vid in list(per.keys()):
        if per[vid]["vendor_name"] == "Vendor":
            vd = await db.vendors.find_one({"id": vid}, {"_id": 0, "name": 1})
            if vd:
                per[vid]["vendor_name"] = vd["name"]
    return list(per.values())

async def _shift_report(shift):
    orders = await db.orders.find({"shift_id": shift["id"], "status": "paid"}, {"_id": 0}).to_list(5000)
    by_type = {"dine_in": 0, "take_away": 0, "retail": 0}
    by_pm = {}
    total = 0
    for o in orders:
        by_type[o["order_type"]] = by_type.get(o["order_type"], 0) + o["total"]
        by_pm[o.get("payment_method_name", "?")] = by_pm.get(o.get("payment_method_name", "?"), 0) + o["total"]
        total += o["total"]
    cash = sum(o["total"] for o in orders if o.get("payment_method_type") == "cash")
    # Kas harian dalam shift (setoran/pengambilan) — digabung ke laporan shift
    moves = await db.cash_movements.find({"shift_id": shift["id"]}, {"_id": 0}).to_list(2000)
    cash_in = sum(m["amount"] for m in moves if m["type"] == "in")
    cash_out = sum(m["amount"] for m in moves if m["type"] == "out")
    vendor_rows = await _vendor_share_from_orders(orders)
    return {"order_count": len(orders), "total_sales": round(total, 2), "by_type": by_type,
            "by_payment": by_pm, "expected_cash": round(shift["opening_cash"] + cash, 2),
            "cash_in": round(cash_in, 2), "cash_out": round(cash_out, 2),
            "cash_net": round(cash_in - cash_out, 2),
            "vendor_share": vendor_rows,
            "vendor_total_share": round(sum(v["share"] for v in vendor_rows), 2)}

@api.get("/shifts/current/vendor")
async def shift_vendor_preview(user: dict = Depends(get_current_user)):
    """Preview bagian vendor dari shift yang sedang berjalan (untuk form penutupan)."""
    shift = await _current_shift(user)
    if not shift:
        raise HTTPException(400, "Tidak ada shift terbuka")
    orders = await db.orders.find({"shift_id": shift["id"], "status": "paid"}, {"_id": 0}).to_list(5000)
    return {"vendors": await _vendor_share_from_orders(orders)}

@api.get("/shifts")
async def list_shifts(admin: dict = Depends(require_admin)):
    return await db.shifts.find({}, {"_id": 0}).sort("opened_at", -1).to_list(200)

# ================================================================== REPORTS
@api.get("/reports/summary")
async def report_summary(date_str: Optional[str] = Query(None, alias="date"),
                         admin: dict = Depends(admin_or_kasir)):
    d = date_str or wib_today()
    start, end = wib_day_range(d)
    q = {"status": "paid", "created_at": {"$gte": start, "$lt": end}}
    orders = await db.orders.find(q, {"_id": 0}).to_list(5000)
    by_type = {"dine_in": {"count": 0, "total": 0}, "take_away": {"count": 0, "total": 0}, "retail": {"count": 0, "total": 0}}
    by_pm = {}
    product_sales = {}
    total = 0
    total_discount = 0
    total_cost = 0
    prods = await db.products.find({}, {"_id": 0, "id": 1, "category_id": 1, "type": 1}).to_list(5000)
    prod_map = {p["id"]: p for p in prods}
    cats_all = await db.categories.find({}, {"_id": 0}).to_list(1000)
    cat_map = {c["id"]: c for c in cats_all}
    cat_sales = {}
    group_totals = {"makanan": 0, "minuman": 0, "retail": 0}
    for o in orders:
        bt = by_type[o["order_type"]]
        bt["count"] += 1
        bt["total"] += o["total"]
        total += o["total"]
        total_discount += o.get("discount", 0)
        by_pm[o.get("payment_method_name", "?")] = by_pm.get(o.get("payment_method_name", "?"), 0) + o["total"]
        for it in o["items"]:
            line = it["price"] * it["qty"]
            ps = product_sales.setdefault(it["name"], {"qty": 0, "total": 0, "cost": 0})
            ps["qty"] += it["qty"]
            ps["total"] += line
            ps["cost"] += it.get("cost", 0) * it["qty"]
            total_cost += it.get("cost", 0) * it["qty"]
            pinfo = prod_map.get(it.get("product_id"), {})
            itype = pinfo.get("type") or it.get("type") or "retail"
            if itype in group_totals:
                group_totals[itype] += line
            cid = pinfo.get("category_id")
            if cid:
                cinfo = cat_map.get(cid, {})
                cs = cat_sales.setdefault(cid, {"category_id": cid, "name": cinfo.get("name", "?"),
                                                "type": cinfo.get("type", itype), "qty": 0, "total": 0})
                cs["qty"] += it["qty"]
                cs["total"] += line
    top = []
    for k, v in sorted(product_sales.items(), key=lambda x: x[1]["total"], reverse=True)[:8]:
        profit = v["total"] - v["cost"]
        margin = round(profit / v["total"] * 100, 1) if v["total"] else 0
        top.append({"name": k, "qty": v["qty"], "total": round(v["total"], 2),
                    "cost": round(v["cost"], 2), "profit": round(profit, 2), "margin": margin})
    fnb_total = by_type["dine_in"]["total"] + by_type["take_away"]["total"]
    by_cat = sorted(cat_sales.values(), key=lambda x: x["total"], reverse=True)
    for c in by_cat:
        c["total"] = round(c["total"], 2)
    category_report = {
        "makanan": {"total": round(group_totals["makanan"], 2),
                    "categories": [c for c in by_cat if c["type"] == "makanan"]},
        "minuman": {"total": round(group_totals["minuman"], 2),
                    "categories": [c for c in by_cat if c["type"] == "minuman"]},
        "retail": {"total": round(group_totals["retail"], 2)},
    }
    low_stock = await db.products.aggregate([
        {"$match": {"track_stock": True, "active": True}},
        {"$addFields": {"eff_min": {"$ifNull": ["$min_stock", LOW_STOCK_THRESHOLD]}}},
        {"$match": {"$expr": {"$lte": ["$stock", "$eff_min"]}}},
        {"$project": {"_id": 0, "name": 1, "sku": 1, "stock": 1, "min_stock": "$eff_min"}},
        {"$sort": {"stock": 1}},
        {"$limit": 100},
    ]).to_list(100)
    cash_moves = await db.cash_movements.find({"created_at": {"$gte": start, "$lt": end}}, {"_id": 0}).to_list(2000)
    cash_in = sum(m["amount"] for m in cash_moves if m["type"] == "in")
    cash_out = sum(m["amount"] for m in cash_moves if m["type"] == "out")
    vendor_summary = await _vendor_report(date_str=d)
    return {"date": d, "vendor_summary": vendor_summary, "total_sales": round(total, 2), "order_count": len(orders),
            "total_discount": round(total_discount, 2), "by_type": by_type, "by_payment": by_pm,
            "fnb_total": round(fnb_total, 2), "retail_total": round(by_type["retail"]["total"], 2),
            "top_products": top, "low_stock": low_stock, "low_stock_threshold": LOW_STOCK_THRESHOLD,
            "total_cost": round(total_cost, 2), "gross_profit": round(total - total_cost, 2),
            "cash_in": round(cash_in, 2), "cash_out": round(cash_out, 2), "cash_net": round(cash_in - cash_out, 2),
            "category_report": category_report}

@api.get("/reports/range")
async def report_range(start: str, end: str, admin: dict = Depends(admin_or_kasir)):
    s_utc, _ = wib_day_range(start)
    _, e_utc = wib_day_range(end)
    q = {"status": "paid", "created_at": {"$gte": s_utc, "$lt": e_utc}}
    orders = await db.orders.find(q, {"_id": 0}).to_list(20000)
    daily = {}
    for o in orders:
        day = wib_day_of(o["created_at"])
        daily.setdefault(day, {"date": day, "total": 0, "count": 0})
        daily[day]["total"] += o["total"]
        daily[day]["count"] += 1
    return {"daily": sorted(daily.values(), key=lambda x: x["date"])}

async def _period_report(start: str, end: str):
    s_utc, _ = wib_day_range(start)
    _, e_utc = wib_day_range(end)
    q = {"status": "paid", "created_at": {"$gte": s_utc, "$lt": e_utc}}
    orders = await db.orders.find(q, {"_id": 0}).to_list(50000)
    prods = await db.products.find({}, {"_id": 0, "id": 1, "category_id": 1, "type": 1}).to_list(5000)
    prod_map = {p["id"]: p for p in prods}
    cats_all = await db.categories.find({}, {"_id": 0}).to_list(1000)
    cat_map = {c["id"]: c for c in cats_all}
    by_type = {"dine_in": {"count": 0, "total": 0}, "take_away": {"count": 0, "total": 0}, "retail": {"count": 0, "total": 0}}
    by_pm = {}
    cat_sales = {}
    group_totals = {"makanan": 0, "minuman": 0, "retail": 0}
    product_sales = {}
    total = 0; total_discount = 0; total_cost = 0
    for o in orders:
        bt = by_type[o["order_type"]]
        bt["count"] += 1
        bt["total"] += o["total"]
        total += o["total"]
        total_discount += o.get("discount", 0)
        by_pm[o.get("payment_method_name", "?")] = by_pm.get(o.get("payment_method_name", "?"), 0) + o["total"]
        for it in o["items"]:
            line = it["price"] * it["qty"]
            ps = product_sales.setdefault(it["name"], {"qty": 0, "total": 0, "cost": 0})
            ps["qty"] += it["qty"]; ps["total"] += line; ps["cost"] += it.get("cost", 0) * it["qty"]
            total_cost += it.get("cost", 0) * it["qty"]
            pinfo = prod_map.get(it.get("product_id"), {})
            itype = pinfo.get("type") or it.get("type") or "retail"
            if itype in group_totals:
                group_totals[itype] += line
            cid = pinfo.get("category_id")
            if cid:
                cinfo = cat_map.get(cid, {})
                cs = cat_sales.setdefault(cid, {"category_id": cid, "name": cinfo.get("name", "?"),
                                                "type": cinfo.get("type", itype), "qty": 0, "total": 0})
                cs["qty"] += it["qty"]; cs["total"] += line
    by_cat = sorted(cat_sales.values(), key=lambda x: x["total"], reverse=True)
    for c in by_cat:
        c["total"] = round(c["total"], 2)
    category_report = {
        "makanan": {"total": round(group_totals["makanan"], 2), "categories": [c for c in by_cat if c["type"] == "makanan"]},
        "minuman": {"total": round(group_totals["minuman"], 2), "categories": [c for c in by_cat if c["type"] == "minuman"]},
        "retail": {"total": round(group_totals["retail"], 2), "categories": [c for c in by_cat if c["type"] == "retail"]},
    }
    top = []
    for k, v in sorted(product_sales.items(), key=lambda x: x[1]["total"], reverse=True)[:10]:
        profit = v["total"] - v["cost"]
        margin = round(profit / v["total"] * 100, 1) if v["total"] else 0
        top.append({"name": k, "qty": v["qty"], "total": round(v["total"], 2), "profit": round(profit, 2), "margin": margin})
    vendor = await _vendor_report(start=start, end=end)
    return {"start": start, "end": end, "total_sales": round(total, 2), "order_count": len(orders),
            "total_discount": round(total_discount, 2), "total_cost": round(total_cost, 2),
            "gross_profit": round(total - total_cost, 2), "by_type": by_type, "by_payment": by_pm,
            "category_report": category_report, "top_products": top, "vendor": vendor}

@api.get("/reports/period")
async def report_period(start: str, end: str, admin: dict = Depends(admin_or_kasir)):
    return await _period_report(start, end)

def _period_report_lines(rep):
    cr = rep["category_report"]
    L = ["*Laporan Grand Aceh Kuliner*", f"Periode: {rep['start']} s/d {rep['end']}", "",
         f"Total Penjualan: Rp{rep['total_sales']:,.0f}",
         f"Jumlah Order: {rep['order_count']}",
         f"Laba Kotor: Rp{rep['gross_profit']:,.0f}",
         f"Total Diskon: Rp{rep['total_discount']:,.0f}", ""]
    for key, label in [("makanan", "Makanan"), ("minuman", "Minuman"), ("retail", "Retail")]:
        g = cr[key]
        L.append(f"{label}: Rp{g['total']:,.0f}")
        for c in g["categories"]:
            L.append(f"  - {c['name']} (x{c['qty']}): Rp{c['total']:,.0f}")
    v = rep.get("vendor", {})
    L += ["", f"Bagi Hasil Vendor: Rp{v.get('total_vendor_share', 0):,.0f} (omzet Rp{v.get('total_gross', 0):,.0f})"]
    for r in v.get("rows", []):
        L.append(f"  - {r['vendor_name']}: Rp{r['vendor_share']:,.0f}")
    return L

@api.get("/reports/period/export/excel")
async def export_period_excel(start: str, end: str, admin: dict = Depends(admin_or_kasir)):
    import openpyxl
    rep = await _period_report(start, end)
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Laporan"
    ws.append(["Laporan Grand Aceh Kuliner"])
    ws.append(["Periode", f"{start} s/d {end}"])
    ws.append([])
    ws.append(["Total Penjualan", rep["total_sales"]])
    ws.append(["Jumlah Order", rep["order_count"]])
    ws.append(["Laba Kotor", rep["gross_profit"]])
    ws.append(["Total Diskon", rep["total_discount"]])
    ws.append([])
    for key, label in [("makanan", "Makanan"), ("minuman", "Minuman"), ("retail", "Retail")]:
        g = rep["category_report"][key]
        ws.append([label, "", g["total"]])
        for c in g["categories"]:
            ws.append(["  " + c["name"], c["qty"], c["total"]])
        ws.append([])
    v = rep.get("vendor", {})
    ws.append(["Bagi Hasil Vendor"])
    ws.append(["Vendor", "Qty", "Omzet", "Bagi Hasil Vendor", "Bagian Outlet"])
    for r in v.get("rows", []):
        ws.append([r["vendor_name"], r["qty"], r["gross"], r["vendor_share"], r["outlet_share"]])
    ws.append(["TOTAL", "", v.get("total_gross", 0), v.get("total_vendor_share", 0), v.get("total_outlet_share", 0)])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename=laporan-{start}_{end}.xlsx"})

@api.get("/reports/period/export/pdf")
async def export_period_pdf(start: str, end: str, admin: dict = Depends(admin_or_kasir)):
    from fpdf import FPDF
    rep = await _period_report(start, end)
    pdf = FPDF(); pdf.add_page(); pdf.set_font("Helvetica", size=12)
    for ln in _period_report_lines(rep):
        txt = ln.replace("*", "").encode("latin-1", "replace").decode("latin-1")
        pdf.multi_cell(pdf.epw, 7, txt or " ")
    out = io.BytesIO(bytes(pdf.output()))
    return StreamingResponse(out, media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename=laporan-{start}_{end}.pdf"})

# ================================================================== INVENTORY (Retail)
class PurchaseIn(BaseModel):
    product_id: str
    qty: int = Field(gt=0)
    unit_cost: float = Field(ge=0)
    note: Optional[str] = ""

@api.post("/purchases")
async def create_purchase(body: PurchaseIn, admin: dict = Depends(admin_or_input)):
    p = await db.products.find_one({"id": body.product_id}, {"_id": 0})
    if not p:
        raise HTTPException(404, "Produk tidak ditemukan")
    if not p.get("track_stock"):
        raise HTTPException(400, "Pembelian stok hanya untuk produk retail")
    await db.products.update_one({"id": body.product_id},
                                 {"$inc": {"stock": body.qty}, "$set": {"cost": body.unit_cost}})
    doc = {"id": new_id(), "product_id": p["id"], "product_name": p["name"], "sku": p["sku"],
           "qty": body.qty, "unit_cost": body.unit_cost, "total_cost": round(body.qty * body.unit_cost, 2),
           "note": body.note, "by": admin["name"], "created_at": now_utc().isoformat()}
    await db.purchases.insert_one(doc)
    doc.pop("_id", None)
    return {**doc, "new_stock": p.get("stock", 0) + body.qty}

class BulkPurchaseItemIn(BaseModel):
    product_id: Optional[str] = None
    create_new: bool = False
    name: Optional[str] = None
    category_id: Optional[str] = None
    price: Optional[float] = None
    qty: int = Field(gt=0)
    unit_cost: float = Field(ge=0)

class BulkPurchaseIn(BaseModel):
    items: List[BulkPurchaseItemIn]
    note: Optional[str] = "Faktur AI"

@api.post("/purchases/bulk")
async def create_purchases_bulk(body: BulkPurchaseIn, admin: dict = Depends(admin_or_input)):
    if not body.items:
        raise HTTPException(400, "Tidak ada item untuk disimpan")
    # Validate all items before writing anything
    for it in body.items:
        if it.create_new:
            if not (it.name and it.name.strip()) or not it.category_id:
                raise HTTPException(400, f"Produk baru '{it.name or '?'}' butuh nama & kategori")
            if not await db.categories.find_one({"id": it.category_id}):
                raise HTTPException(400, f"Kategori untuk '{it.name}' tidak valid")
        else:
            if not it.product_id:
                raise HTTPException(400, "Item lama butuh product_id")
            p = await db.products.find_one({"id": it.product_id}, {"_id": 0})
            if not p:
                raise HTTPException(404, f"Produk '{it.name or it.product_id}' tidak ditemukan")
            if not p.get("track_stock"):
                raise HTTPException(400, f"'{p['name']}' bukan produk retail (tidak melacak stok)")
    saved, created_products = [], 0
    for it in body.items:
        if it.create_new:
            prod = {"id": new_id(), "name": it.name.strip(), "sku": "AI-" + new_id()[:6],
                    "category_id": it.category_id, "type": "retail",
                    "price": float(it.price if it.price is not None else it.unit_cost),
                    "cost": float(it.unit_cost), "description": "", "image": "", "active": True,
                    "sold_out": False, "stock": 0, "min_stock": 10, "track_stock": True,
                    "created_at": now_utc().isoformat()}
            await db.products.insert_one(prod)
            pid, pname, psku, base_stock = prod["id"], prod["name"], prod["sku"], 0
            created_products += 1
        else:
            p = await db.products.find_one({"id": it.product_id}, {"_id": 0})
            pid, pname, psku, base_stock = p["id"], p["name"], p["sku"], p.get("stock", 0)
        await db.products.update_one({"id": pid}, {"$inc": {"stock": it.qty}, "$set": {"cost": it.unit_cost}})
        doc = {"id": new_id(), "product_id": pid, "product_name": pname, "sku": psku,
               "qty": it.qty, "unit_cost": it.unit_cost, "total_cost": round(it.qty * it.unit_cost, 2),
               "note": body.note, "by": admin["name"], "created_at": now_utc().isoformat()}
        await db.purchases.insert_one(doc)
        doc.pop("_id", None)
        saved.append({**doc, "new_stock": base_stock + it.qty})
    return {"saved": len(saved), "created_products": created_products,
            "total_cost": round(sum(s["total_cost"] for s in saved), 2), "items": saved}

@api.get("/purchases")
async def list_purchases(date_str: Optional[str] = Query(None, alias="date"), admin: dict = Depends(admin_or_input)):
    q = {}
    if date_str:
        s, e = wib_day_range(date_str)
        q["created_at"] = {"$gte": s, "$lt": e}
    return await db.purchases.find(q, {"_id": 0}).sort("created_at", -1).to_list(1000)

class OpnameIn(BaseModel):
    product_id: str
    counted_stock: int = Field(ge=0)
    note: Optional[str] = ""

@api.post("/stock-opname")
async def create_opname(body: OpnameIn, admin: dict = Depends(admin_or_input)):
    p = await db.products.find_one({"id": body.product_id}, {"_id": 0})
    if not p:
        raise HTTPException(404, "Produk tidak ditemukan")
    if not p.get("track_stock"):
        raise HTTPException(400, "Stok opname hanya untuk produk retail")
    system_stock = p.get("stock", 0)
    diff = body.counted_stock - system_stock
    await db.products.update_one({"id": body.product_id}, {"$set": {"stock": body.counted_stock}})
    doc = {"id": new_id(), "product_id": p["id"], "product_name": p["name"], "sku": p["sku"],
           "system_stock": system_stock, "counted_stock": body.counted_stock, "difference": diff,
           "note": body.note, "by": admin["name"], "created_at": now_utc().isoformat()}
    await db.stock_opname.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.get("/stock-opname")
async def list_opname(date_str: Optional[str] = Query(None, alias="date"), admin: dict = Depends(admin_or_input)):
    q = {}
    if date_str:
        s, e = wib_day_range(date_str)
        q["created_at"] = {"$gte": s, "$lt": e}
    return await db.stock_opname.find(q, {"_id": 0}).sort("created_at", -1).to_list(1000)

# ================================================================== CASH MOVEMENTS
class CashIn(BaseModel):
    type: Literal["in", "out"]
    amount: float = Field(gt=0)
    category: str = "Lainnya"
    note: Optional[str] = ""

@api.post("/cash")
async def create_cash(body: CashIn, user: dict = Depends(get_current_user)):
    shift = await _current_shift(user)
    doc = {"id": new_id(), "type": body.type, "amount": body.amount, "category": body.category,
           "note": body.note, "cashier_id": user["id"], "cashier_name": user["name"],
           "shift_id": shift["id"] if shift else None, "created_at": now_utc().isoformat()}
    await db.cash_movements.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.get("/cash")
async def list_cash(date_str: Optional[str] = Query(None, alias="date"), user: dict = Depends(get_current_user)):
    d = date_str or wib_today()
    s, e = wib_day_range(d)
    moves = await db.cash_movements.find({"created_at": {"$gte": s, "$lt": e}}, {"_id": 0}).sort("created_at", -1).to_list(2000)
    cin = sum(m["amount"] for m in moves if m["type"] == "in")
    cout = sum(m["amount"] for m in moves if m["type"] == "out")
    return {"date": d, "movements": moves, "cash_in": round(cin, 2), "cash_out": round(cout, 2),
            "cash_net": round(cin - cout, 2)}

# ================================================================== SYNC (local outlet server <-> cloud)
class SyncPushIn(BaseModel):
    orders: List[OrderIn]

@api.get("/sync/master")
async def sync_master(user: dict = Depends(get_current_user)):
    """Pull master data snapshot for a local outlet server / offline client."""
    products = await db.products.find({}, {"_id": 0}).to_list(5000)
    categories = await db.categories.find({}, {"_id": 0}).to_list(500)
    tables = await db.tables.find({"deleted": {"$ne": True}}, {"_id": 0}).to_list(500)
    pms = await db.payment_methods.find({}, {"_id": 0}).to_list(100)
    return {"server_time": now_utc().isoformat(), "products": products, "categories": categories,
            "tables": tables, "payment_methods": pms}

@api.post("/sync/push")
async def sync_push(body: SyncPushIn, user: dict = Depends(get_current_user)):
    """Push a batch of offline orders to cloud. Idempotent via client_ref."""
    results = []
    synced = 0
    for o in body.orders:
        try:
            res = await create_order(o, user)
            synced += 1
            results.append({"client_ref": o.client_ref, "status": "ok", "order_number": res.get("order_number")})
        except HTTPException as e:
            results.append({"client_ref": o.client_ref, "status": "error", "detail": e.detail})
    return {"synced": synced, "total": len(body.orders), "results": results, "server_time": now_utc().isoformat()}

# ================================================================== AI
def _get_chat(session, system, model):
    if not EMERGENT_LLM_KEY:
        raise HTTPException(400, "AI belum dikonfigurasi. Isi API key & endpoint provider Anda di Pengaturan AI.")
    from emergentintegrations.llm.chat import LlmChat
    return LlmChat(api_key=EMERGENT_LLM_KEY, session_id=session, system_message=system).with_model("gemini", model)

AI_FEATURES = {"description": "Deskripsi Produk", "image": "Gambar Produk", "summary": "Analisis Laporan", "vision": "Baca Faktur (Vision)", "assistant": "Asisten Admin"}

async def _ai_cfg(feature="description"):
    """Per-feature AI provider config: DB settings (editable in UI) override .env.
    Fallback order: feature-specific -> legacy flat -> .env defaults."""
    doc = await db.settings.find_one({"_id": "ai"}) or {}
    feat = (doc.get("features", {}) or {}).get(feature, {}) or {}
    return {
        "base_url": feat.get("base_url") or doc.get("openai_base_url") or OPENAI_COMPAT_BASE_URL,
        "api_key": feat.get("api_key") or doc.get("openai_api_key") or OPENAI_COMPAT_API_KEY,
        "model": feat.get("model") or doc.get("openai_model") or OPENAI_COMPAT_MODEL,
    }

async def _ai_provider_cfg(feature="description"):
    """Resolve which provider ('gemini' | 'chenzk') to use for a feature and its creds."""
    doc = await db.settings.find_one({"_id": "ai"}) or {}
    feat = (doc.get("features", {}) or {}).get(feature, {}) or {}
    provider = feat.get("provider")
    if not provider:
        # Infer for backward-compat: OpenAI-compatible creds present -> chenzk, else gemini.
        if feat.get("api_key") and (feat.get("base_url") or OPENAI_COMPAT_BASE_URL):
            provider = "chenzk"
        else:
            provider = "gemini"
    if provider == "gemini":
        return {"provider": "gemini",
                "api_key": feat.get("api_key") or GEMINI_API_KEY,
                "model": feat.get("model") or GEMINI_TEXT_MODEL}
    return {"provider": "chenzk",
            "base_url": (feat.get("base_url") or doc.get("openai_base_url") or OPENAI_COMPAT_BASE_URL or CHENZK_BASE_URL),
            "api_key": feat.get("api_key") or doc.get("openai_api_key") or OPENAI_COMPAT_API_KEY,
            "model": feat.get("model") or doc.get("openai_model") or OPENAI_COMPAT_MODEL}

async def _ai_chat(messages, system="", feature="description", temperature=0.5, max_tokens=4000):
    """Unified text chat across providers. `messages`=[{role:'user'|'assistant',content:str}].
    Routes to Gemini REST (X-goog-api-key) or chenzk (OpenAI-compatible), else Emergent fallback."""
    cfg = await _ai_provider_cfg(feature)
    if cfg["provider"] == "gemini" and cfg.get("api_key"):
        import httpx
        model = cfg["model"] or GEMINI_TEXT_MODEL
        contents = []
        for m in messages:
            role = "model" if m.get("role") == "assistant" else "user"
            contents.append({"role": role, "parts": [{"text": str(m.get("content", ""))}]})
        payload = {"contents": contents,
                   "generationConfig": {"temperature": temperature, "maxOutputTokens": max_tokens}}
        if system:
            payload["systemInstruction"] = {"parts": [{"text": system}]}
        url = f"{GEMINI_REST_URL}/{model}:generateContent"
        try:
            async with httpx.AsyncClient(timeout=60) as c:
                r = await c.post(url, headers={"Content-Type": "application/json", "X-goog-api-key": cfg["api_key"]}, json=payload)
        except Exception as e:
            raise HTTPException(400, f"Gagal menghubungi Gemini: {e}")
        if r.status_code != 200:
            raise HTTPException(400, f"Gemini menolak permintaan (HTTP {r.status_code}): {r.text[:200]}")
        data = r.json()
        try:
            parts = data["candidates"][0]["content"]["parts"]
            return "".join(p.get("text", "") for p in parts).strip()
        except Exception:
            raise HTTPException(400, "Gemini tidak mengembalikan teks. Coba lagi atau ganti model.")
    if cfg["provider"] == "chenzk" and cfg.get("api_key") and cfg.get("base_url"):
        from openai import OpenAI
        msgs = ([{"role": "system", "content": system}] if system else []) + \
               [{"role": ("assistant" if m.get("role") == "assistant" else "user"), "content": str(m.get("content", ""))} for m in messages]

        def run():
            client = OpenAI(api_key=cfg["api_key"], base_url=cfg["base_url"])
            r = client.chat.completions.create(model=cfg["model"], messages=msgs,
                                               temperature=temperature, max_tokens=max_tokens)
            msg = r.choices[0].message
            content = (msg.content or "").strip()
            if not content:
                content = (getattr(msg, "reasoning_content", "") or "").strip()
            return content
        return await asyncio.to_thread(run)
    # Fallback: Emergent universal key (single-turn only)
    if not EMERGENT_LLM_KEY:
        raise HTTPException(400, "AI belum dikonfigurasi. Pilih provider (Gemini/chenzk) dan isi API key di Pengaturan AI.")
    from emergentintegrations.llm.chat import UserMessage
    chat = _get_chat(new_id(), system, "gemini-2.5-flash")
    last = messages[-1]["content"] if messages else ""
    return (await chat.send_message(UserMessage(text=last))).strip()

async def _gemini_text(system, prompt, feature="description"):
    """Text generation for description/summary — routes via unified provider chat."""
    return await _ai_chat([{"role": "user", "content": prompt}], system=system, feature=feature,
                          temperature=0.5, max_tokens=800 if feature == "description" else 4000)

async def _gemini_image(prompt):
    """Image via OpenAI-compatible image endpoint (only when explicitly configured for 'image'),
    else user's own Gemini key, else Emergent universal key."""
    doc = await db.settings.find_one({"_id": "ai"}) or {}
    feat = (doc.get("features", {}) or {}).get("image", {}) or {}
    if feat.get("api_key") and feat.get("base_url") and feat.get("model"):
        from openai import OpenAI

        def run():
            client = OpenAI(api_key=feat["api_key"], base_url=feat["base_url"])
            r = client.images.generate(model=feat["model"], prompt=prompt, n=1, size="1024x1024")
            d = r.data[0]
            b64 = getattr(d, "b64_json", None)
            if b64:
                return f"data:image/png;base64,{b64}"
            return getattr(d, "url", None)
        return await asyncio.to_thread(run)
    if GEMINI_API_KEY:
        from google import genai
        from google.genai import types
        import base64

        def run():
            client = genai.Client(api_key=GEMINI_API_KEY)
            r = client.models.generate_content(
                model=GEMINI_IMAGE_MODEL, contents=prompt,
                config=types.GenerateContentConfig(response_modalities=["IMAGE", "TEXT"]),
            )
            for cand in (r.candidates or []):
                for part in (cand.content.parts or []):
                    inline = getattr(part, "inline_data", None)
                    if inline and inline.data:
                        data = inline.data
                        b64 = base64.b64encode(data).decode() if isinstance(data, (bytes, bytearray)) else data
                        return f"data:{inline.mime_type or 'image/png'};base64,{b64}"
            return None
        return await asyncio.to_thread(run)
    if not EMERGENT_LLM_KEY:
        raise HTTPException(400, "Generator gambar AI belum dikonfigurasi. Isi provider gambar Anda di Pengaturan AI, atau unggah gambar manual.")
    from emergentintegrations.llm.chat import UserMessage
    chat = _get_chat(new_id(), "You are a professional food & product photographer.",
                     "gemini-3.1-flash-image-preview").with_params(modalities=["image", "text"])
    _, images = await chat.send_message_multimodal_response(UserMessage(text=prompt))
    if images:
        img = images[0]
        return f"data:{img['mime_type']};base64,{img['data']}"
    return None

async def _save_image_local(src: str) -> str:
    """Persist an AI image (base64 data URL or remote URL) to local disk; return stable /api/uploads path so it never expires."""
    if not src:
        return src

    def run():
        import base64 as _b64, urllib.request
        if src.startswith("data:"):
            header, _, b64data = src.partition(",")
            mime = header[5:].split(";")[0] or "image/png"
            raw = _b64.b64decode(b64data)
        elif src.startswith("http"):
            req = urllib.request.Request(src, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=30) as resp:
                raw = resp.read()
                mime = resp.headers.get_content_type() or "image/png"
        else:
            return src
        ext = {"image/png": "png", "image/jpeg": "jpg", "image/jpg": "jpg", "image/webp": "webp"}.get(mime, "png")
        fname = f"{new_id()}.{ext}"
        (UPLOAD_DIR / fname).write_bytes(raw)
        return f"/api/uploads/{fname}"
    return await asyncio.to_thread(run)

@api.get("/health")
async def health():
    return {"app": "gak-pos", "ok": True}

@api.get("/ota/version")
async def ota_version():
    """Versi OTA yang disajikan server — lewat API (CORS FastAPI *), jadi APK
    tidak perlu fetch /ota/version.json (nginx) yang rawan masalah CORS.
    Membaca file versi dari mount /host-project (ro)."""
    import json as _json
    version = ""
    for base in ("/host-project", os.environ.get("HOST_PROJECT_DIR", "")):
        if not base:
            continue
        p = os.path.join(base, "frontend", "build", "ota", "version.json")
        try:
            with open(p) as f:
                d = _json.load(f)
                version = str(d.get("version", "") or "").strip()
            if version:
                break
        except Exception:
            continue
    return {"version": version, "url": "/ota/bundle.zip"}

@api.get("/uploads/{fname}")
async def get_upload(fname: str):
    fp = UPLOAD_DIR / os.path.basename(fname)
    if not fp.exists():
        raise HTTPException(404, "File tidak ditemukan")
    return FileResponse(str(fp))

@api.get("/installers/project-zip")
async def download_project_zip(admin: dict = Depends(require_admin)):
    if not (PROJECT_ROOT / "docker-compose.yml").exists():
        raise HTTPException(404, "Folder proyek tidak tersedia di lingkungan ini.")
    EXCLUDE_DIRS = {"node_modules", ".git", "build", "__pycache__", ".venv", "venv",
                    "uploads", "backups", ".wwebjs_auth", ".wwebjs_cache", ".emergent",
                    "dist", ".pytest_cache", "test_reports", "memory", ".gradle",
                    ".idea", ".vscode", "coverage"}
    EXCLUDE_FILES = {".env", ".env.docker", ".env.local", ".DS_Store"}

    def build():
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED, compresslevel=1) as zf:
            for root, dirs, files in os.walk(PROJECT_ROOT):
                dirs[:] = [d for d in dirs if d not in EXCLUDE_DIRS]
                for f in files:
                    if f in EXCLUDE_FILES or f.endswith(".env.local"):
                        continue
                    fp = os.path.join(root, f)
                    if os.path.islink(fp) or not os.path.isfile(fp):
                        continue
                    rel = os.path.relpath(fp, PROJECT_ROOT)
                    zf.write(fp, os.path.join("grand-aceh-pos", rel))
        buf.seek(0)
        return buf.read()

    data = await asyncio.to_thread(build)
    return Response(content=data, media_type="application/zip",
                    headers={"Content-Disposition": "attachment; filename=grand-aceh-pos.zip"})

@api.get("/backup/export")
async def backup_export(admin: dict = Depends(require_admin)):
    names = await db.list_collection_names()
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for name in names:
            if name.startswith("system."):
                continue
            docs = await db[name].find({}).to_list(200000)
            zf.writestr(f"{name}.json", json_util.dumps(docs))
    buf.seek(0)
    ts = now_utc().strftime("%Y%m%d-%H%M%S")
    return Response(content=buf.getvalue(), media_type="application/zip",
                    headers={"Content-Disposition": f"attachment; filename=gak-backup-{ts}.zip"})

@api.post("/backup/import")
async def backup_import(file: UploadFile = File(...), admin: dict = Depends(require_admin)):
    raw = await file.read()
    try:
        zf = zipfile.ZipFile(io.BytesIO(raw))
    except Exception:
        raise HTTPException(400, "File backup tidak valid (.zip)")
    restored = 0
    for nm in zf.namelist():
        if not nm.endswith(".json"):
            continue
        coll = nm[:-5]
        docs = json_util.loads(zf.read(nm).decode("utf-8"))
        await db[coll].delete_many({})
        if docs:
            await db[coll].insert_many(docs)
        restored += 1
    return {"restored_collections": restored}

DOCKER_SOCK = "/var/run/docker.sock"

def _update_enabled():
    return os.path.exists(DOCKER_SOCK) and bool(os.environ.get("HOST_PROJECT_DIR"))

@api.get("/admin/update/status")
async def update_status(admin: dict = Depends(require_admin)):
    running = False
    log = ""
    if _update_enabled():
        try:
            import docker
            cli = docker.from_env()
            try:
                c = cli.containers.get("gak-updater")
                c.reload()
                running = c.status == "running"
                log = c.logs(tail=20).decode("utf-8", "ignore")[-1800:]
            except Exception:
                running = False
        except Exception:
            pass
    return {"enabled": _update_enabled(), "running": running, "log": log,
            "host_project_dir": os.environ.get("HOST_PROJECT_DIR")}

VIBE_UPDATE_BASE_URL = os.environ.get("VIBE_UPDATE_BASE_URL", "https://taqim258.vibecoder.co.id/pos-grand-update")
VIBE_REPORT_URL = os.environ.get("VIBE_REPORT_URL", "https://taqim258.vibecoder.co.id/pos-grand-update/rpt.php")
VIBE_REPORT_TOKEN = os.environ.get("VIBE_REPORT_TOKEN", "gak_rpt_7f3c9e1b")
VIBE_BACKUP_URL = os.environ.get("VIBE_BACKUP_URL", "https://taqim258.vibecoder.co.id/pos-grand-update/bkp.php")
VIBE_BACKUP_TOKEN = os.environ.get("VIBE_BACKUP_TOKEN", "gak_bkp_2a8d51c4")
VIBE_FEATURE_URL = os.environ.get("VIBE_FEATURE_URL", "https://taqim258.vibecoder.co.id/pos-grand-update/feat.php")
VIBE_FEATURE_TOKEN = os.environ.get("VIBE_FEATURE_TOKEN", "gak_feat_5b2d9e77")

class FeatureRequestIn(BaseModel):
    message: str
    context: Optional[str] = None

@api.post("/feature-request/send")
async def feature_request_send(body: FeatureRequestIn, admin: dict = Depends(admin_or_kasir)):
    """Kirim permintaan fitur dari tombol 'Usulkan Fitur' (Asisten AI) ke pusat vibecoder.co.id."""
    import urllib.request, json as _json
    msg = (body.message or "").strip()
    if not msg:
        raise HTTPException(400, "Permintaan fitur kosong")
    if len(msg) > 200_000:
        raise HTTPException(400, "Permintaan fitur terlalu panjang")
    ctx = (body.context or "").strip()
    payload = _json.dumps({"ts": datetime.now(timezone.utc).isoformat(), "message": msg, "context": ctx}).encode("utf-8")
    req = urllib.request.Request(
        VIBE_FEATURE_URL, data=payload,
        headers={"Content-Type": "application/json", "X-Gak-Token": VIBE_FEATURE_TOKEN},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as r:
            resp = r.read().decode("utf-8", "ignore")
        return {"ok": True, "response": resp[:500]}
    except Exception as e:
        raise HTTPException(502, f"Gagal mengirim ke vibecoder.co.id: {e}")

class DiagSendIn(BaseModel):
    report: str

@api.post("/diag/send")
async def diag_send(body: DiagSendIn, admin: dict = Depends(require_admin)):
    """Kirim laporan diagnostik dari tombol 'Kirim ke VibeCoder' ke pusat vibecoder.co.id."""
    import urllib.request, json as _json
    if not body.report or len(body.report) > 200_000:
        raise HTTPException(400, "Laporan kosong atau terlalu besar")
    payload = _json.dumps({"ts": datetime.now(timezone.utc).isoformat(), "report": body.report}).encode("utf-8")
    req = urllib.request.Request(
        VIBE_REPORT_URL, data=payload,
        headers={"Content-Type": "application/json", "X-Gak-Token": VIBE_REPORT_TOKEN},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as r:
            resp = r.read().decode("utf-8", "ignore")
        return {"ok": True, "response": resp[:500]}
    except Exception as e:
        raise HTTPException(502, f"Gagal mengirim ke vibecoder.co.id: {e}")

@api.get("/update/check")
async def update_check(admin: dict = Depends(require_admin)):
    """Cek versi terbaru di vibecoder.co.id (dipakai banner 'Versi baru tersedia' & laporan Diagnostik)."""
    import urllib.request, json
    current = ""
    # Folder proyek di-mount ke /host-project (ro) sejak docker-compose diperbarui;
    # fallback ke HOST_PROJECT_DIR untuk kompatibilitas.
    for base in ("/host-project", os.environ.get("HOST_PROJECT_DIR")):
        if not base:
            continue
        try:
            with open(os.path.join(base, ".vibecoder-version"), "r") as f:
                current = f.read().strip()
            if current:
                break
        except Exception:
            continue
    latest = ""
    reachable = False
    try:
        with urllib.request.urlopen(f"{VIBE_UPDATE_BASE_URL}/version.json", timeout=8) as r:
            data = json.loads(r.read().decode("utf-8", "ignore"))
            latest = str(data.get("version", "") or "").strip()
            reachable = True
    except Exception:
        pass
    return {
        "enabled": bool(current),
        "current": current,
        "latest": latest,
        "updateAvailable": bool(current and latest and latest != current),
        "updateCenterReachable": reachable,
        "baseUrl": VIBE_UPDATE_BASE_URL,
    }

@api.post("/backup/send-to-vibecoder")
async def backup_send_to_vibecoder(admin: dict = Depends(require_admin)):
    """Buat backup database lalu kirim salinannya ke pusat vibecoder.co.id (mirror tambahan).

    Menjalankan container docker:cli yang memanggil ./backup-to-vibecoder.sh di folder host.
    Backup lokal tetap dibuat di backups/; salinan di vibecoder bersifat cadangan tambahan.
    """
    if not os.path.exists(DOCKER_SOCK):
        raise HTTPException(400, "Fitur belum aktif. Jalankan update manual sekali (cd ~/grand-aceh-pos && ./update-pi.sh) untuk mengaktifkannya.")
    host_dir = os.environ.get("HOST_PROJECT_DIR")
    if not host_dir:
        raise HTTPException(400, "HOST_PROJECT_DIR belum diset. Jalankan update manual sekali untuk mengaktifkan fitur ini.")
    try:
        import docker
        cli = docker.from_env()
    except Exception as e:
        raise HTTPException(500, f"Docker tidak tersedia dari aplikasi: {e}")
    try:
        for c in cli.containers.list(all=True, filters={"name": "gak-backup-sender"}):
            try:
                c.remove(force=True)
            except Exception:
                pass
        image = os.environ.get("UPDATER_IMAGE", "docker:cli")
        cmd = "apk add --no-cache curl openssl >/dev/null 2>&1; cd /project && ./backup-to-vibecoder.sh"
        cli.containers.run(
            image,
            command=["sh", "-c", cmd],
            detach=True, remove=True, name="gak-backup-sender",
            volumes={
                DOCKER_SOCK: {"bind": DOCKER_SOCK, "mode": "rw"},
                host_dir: {"bind": "/project", "mode": "rw"},
            },
            working_dir="/project",
            environment={
                "VIBE_BACKUP_TOKEN": VIBE_BACKUP_TOKEN,
                "VIBE_BACKUP_PASS": os.environ.get("VIBE_BACKUP_PASS", ""),
                "VIBE_BACKUP_URL": VIBE_BACKUP_URL,
            },
        )
    except Exception as e:
        raise HTTPException(500, f"Gagal memulai backup: {e}")
    return {"started": True, "message": "Backup dibuat & dikirim ke vibecoder.co.id. Cek folder backups/ di server."}

@api.post("/admin/update")
async def admin_update(admin: dict = Depends(require_admin)):
    if not os.path.exists(DOCKER_SOCK):
        raise HTTPException(400, "Fitur update 1-klik belum aktif. Jalankan update manual sekali (cd ~/grand-aceh-pos && ./update-pi.sh) untuk mengaktifkannya.")
    host_dir = os.environ.get("HOST_PROJECT_DIR")
    if not host_dir:
        raise HTTPException(400, "HOST_PROJECT_DIR belum diset. Jalankan update manual sekali untuk mengaktifkan fitur ini.")
    try:
        import docker
        cli = docker.from_env()
    except Exception as e:
        raise HTTPException(500, f"Docker tidak tersedia dari aplikasi: {e}")
    try:
        for c in cli.containers.list(all=True, filters={"name": "gak-updater"}):
            try:
                c.remove(force=True)
            except Exception:
                pass
        image = os.environ.get("UPDATER_IMAGE", "docker:cli")
        cmd = (
            "apk add --no-cache git curl >/dev/null 2>&1; "
            "if [ -f /project/.vibecoder-version ]; then "
            "VER=\"$(cat /project/.vibecoder-version 2>/dev/null || true)\"; "
            "REMOTE=\"$(curl -fsSL -m 20 https://taqim258.vibecoder.co.id/pos-grand-update/version.json | sed -n 's/.*\"version\"[[:space:]]*:[[:space:]]*\"\([^\"]*\)\".*/\1/p' | head -1 || true)\"; "
            "if [ -n \"$REMOTE\" ] && [ \"$REMOTE\" != \"$VER\" ]; then "
            "curl -fsSL -m 300 -o /tmp/gak-pos-update.tar.gz https://taqim258.vibecoder.co.id/pos-grand-update/pos-grand.tar.gz && "
            "tar xzf /tmp/gak-pos-update.tar.gz -C /project && rm -f /tmp/gak-pos-update.tar.gz && "
            "echo \"$REMOTE\" > /project/.vibecoder-version && echo \"Update ke versi $REMOTE\"; "
            "else "
            "echo 'Sudah versi terbaru, lewati unduhan.'; "
            "fi; "
            "else "
            "git config --global --add safe.directory /project; git -C /project pull --ff-only; "
            "fi; "
            "cd /project && docker compose up -d --build"
        )
        cli.containers.run(
            image,
            command=["sh", "-c", cmd],
            detach=True, remove=True, name="gak-updater",
            volumes={
                DOCKER_SOCK: {"bind": DOCKER_SOCK, "mode": "rw"},
                host_dir: {"bind": "/project", "mode": "rw"},
            },
            working_dir="/project",
        )
    except Exception as e:
        raise HTTPException(500, f"Gagal memulai update: {e}")
    return {"started": True, "message": "Update dimulai. Tunggu 2-10 menit, lalu muat ulang halaman."}


class AISettingsIn(BaseModel):
    feature: Literal["description", "image", "summary", "vision", "assistant"]
    provider: Optional[Literal["gemini", "chenzk"]] = None
    base_url: Optional[str] = None
    api_key: Optional[str] = None
    model: Optional[str] = None

def _mask_key(k):
    if not k:
        return ""
    return "••••" + k[-4:] if len(k) >= 4 else "••••"

@api.get("/settings/ai")
async def get_ai_settings(admin: dict = Depends(admin_or_kasir)):
    doc = await db.settings.find_one({"_id": "ai"}) or {}
    feats = doc.get("features", {}) or {}
    out = {}
    for key, label in AI_FEATURES.items():
        f = feats.get(key, {}) or {}
        provider = f.get("provider")
        if not provider:
            provider = "chenzk" if (f.get("api_key") and (f.get("base_url") or OPENAI_COMPAT_BASE_URL)) else "gemini"
        if provider == "gemini":
            akey = f.get("api_key") or GEMINI_API_KEY
            base = ""
            model = f.get("model") or (GEMINI_IMAGE_MODEL if key == "image" else GEMINI_TEXT_MODEL)
        else:  # chenzk / OpenAI-compatible
            if key == "image":
                akey = f.get("api_key")
                base = f.get("base_url") or CHENZK_BASE_URL
                model = f.get("model") or ""
            else:
                akey = f.get("api_key") or doc.get("openai_api_key") or OPENAI_COMPAT_API_KEY
                base = f.get("base_url") or doc.get("openai_base_url") or OPENAI_COMPAT_BASE_URL or CHENZK_BASE_URL
                model = f.get("model") or doc.get("openai_model") or (OPENAI_COMPAT_MODEL if akey else "") or ""
        out[key] = {"label": label, "provider": provider, "base_url": base, "model": model,
                    "api_key_set": bool(akey), "api_key_last4": _mask_key(akey)}
    return {"features": out}

@api.put("/settings/ai")
async def put_ai_settings(body: AISettingsIn, admin: dict = Depends(require_admin)):
    upd = {}
    if body.provider is not None:
        upd[f"features.{body.feature}.provider"] = body.provider
    if body.base_url is not None:
        upd[f"features.{body.feature}.base_url"] = body.base_url.strip()
    if body.model is not None:
        upd[f"features.{body.feature}.model"] = body.model.strip()
    if body.api_key:  # only overwrite key when a new one is provided
        upd[f"features.{body.feature}.api_key"] = body.api_key.strip()
    if upd:
        await db.settings.update_one({"_id": "ai"}, {"$set": upd}, upsert=True)
    return {"ok": True}

def _model_price_rank(mid):
    m = (mid or "").lower()
    cheap = any(k in m for k in ["flash", "mini", "nano", "lite", "micro", "haiku", "small", "8b", "free"])
    return (0 if cheap else 1, m)

@api.get("/settings/ai/models")
async def ai_models(feature: str = "description", admin: dict = Depends(require_admin)):
    import httpx
    cfg = await _ai_cfg(feature)
    if not (cfg["api_key"] and cfg["base_url"]):
        raise HTTPException(400, "Isi & SIMPAN Base URL + API Key dulu, lalu muat model.")
    base = cfg["base_url"].rstrip("/")
    try:
        async with httpx.AsyncClient(timeout=20) as c:
            r = await c.get(f"{base}/models", headers={"Authorization": f"Bearer {cfg['api_key']}"})
    except Exception as e:
        raise HTTPException(400, f"Gagal menghubungi provider: {e}")
    if r.status_code != 200:
        raise HTTPException(400, f"Provider menolak permintaan (HTTP {r.status_code}).")
    data = r.json()
    items = data.get("data", data) if isinstance(data, dict) else data
    ids = [m.get("id") for m in items if isinstance(m, dict) and m.get("id")]
    ids = sorted(set(ids), key=_model_price_rank)
    return {"models": ids}

@api.get("/settings/ai/credit")
async def ai_credit(feature: str = "description", admin: dict = Depends(require_admin)):
    """Best-effort remaining credit lookup via OpenAI-compatible billing endpoints."""
    import httpx, datetime as _dt
    cfg = await _ai_cfg(feature)
    if not (cfg["api_key"] and cfg["base_url"]):
        return {"available": False, "message": "Konfigurasi belum lengkap untuk fitur ini."}
    base = cfg["base_url"].rstrip("/")
    headers = {"Authorization": f"Bearer {cfg['api_key']}"}
    try:
        async with httpx.AsyncClient(timeout=15) as client:
            sub = await client.get(f"{base}/dashboard/billing/subscription", headers=headers)
            if sub.status_code != 200:
                return {"available": False, "message": f"Provider tidak menyediakan info kredit (HTTP {sub.status_code})."}
            s = sub.json()
            total = s.get("hard_limit_usd") or s.get("system_hard_limit_usd") or 0
            used = 0.0
            try:
                end = _dt.date.today() + _dt.timedelta(days=1)
                start = end - _dt.timedelta(days=100)
                u = await client.get(f"{base}/dashboard/billing/usage", headers=headers,
                                     params={"start_date": str(start), "end_date": str(end)})
                if u.status_code == 200:
                    used = (u.json().get("total_usage") or 0) / 100.0
            except Exception:
                pass
            return {"available": True, "total": round(float(total), 2), "used": round(used, 2),
                    "remaining": round(float(total) - used, 2), "currency": "USD"}
    except Exception as e:
        return {"available": False, "message": f"Gagal cek kredit: {e}"}

# ---------------------------------------------------------------- AI ADMIN ASSISTANT
class AIAssistantChatIn(BaseModel):
    session_id: Optional[str] = None
    message: str

class AIAssistantApplyIn(BaseModel):
    action: dict

ASSISTANT_SYSTEM = (
    "Anda adalah 'Asisten AI' untuk aplikasi kasir/POS 'Grand Aceh Kuliner'. "
    "Bahasa jawaban: Indonesia, ringkas, ramah, dan praktis. "
    "Anda bisa menjawab SEMUA pertanyaan: (a) laporan penjualan/belanja (gunakan data laporan_penjualan, "
    "tampilkan Rupiah, jujur bila data tidak tersedia), (b) cara pakai fitur aplikasi (gunakan panduan_aplikasi), "
    "(c) pengetahuan umum usaha. "
    "Anda membantu admin mengelola: produk, kategori, vendor, harga, diskon, dan metode pembayaran. "
    "PENTING: yang dapat MENERAPKAN perubahan data hanyalah admin; jika pengguna berperan kasir, "
    "tetap jawab dengan baik tapi JANGAN sertakan blok aksi — cukup jelaskan bahwa perubahan data perlu admin. "
    "Jika admin hanya bertanya/minta saran, jawab biasa TANPA blok aksi. "
    "Jika admin meminta PERUBAHAN DATA, jelaskan singkat lalu sertakan TEPAT SATU blok aksi berformat: "
    "<ACTION>{\"type\":\"...\", ...}</ACTION> berisi JSON valid (tanpa komentar). "
    "Jenis aksi yang didukung beserta field-nya:\n"
    "1) create_category: {\"type\":\"create_category\",\"name\":str,\"kind\":\"makanan|minuman|retail|vendor\"}\n"
    "2) create_vendor: {\"type\":\"create_vendor\",\"name\":str,\"contact\":str?,\"note\":str?}\n"
    "3) create_payment_method: {\"type\":\"create_payment_method\",\"name\":str,\"pm_type\":\"cash|qris|card\"}\n"
    "4) create_product: {\"type\":\"create_product\",\"name\":str,\"price\":number,\"kind\":\"makanan|minuman|retail|vendor\",\"category_name\":str,\"cost\":number?,\"stock\":number?,\"sku\":str?,\"description\":str?,\"vendor_name\":str?}\n"
    "5) create_products_bulk (BANYAK produk sekaligus dari daftar tempel): {\"type\":\"create_products_bulk\",\"items\":[{ ...field sama seperti create_product... }]}\n"
    "6) update_product: {\"type\":\"update_product\",\"name\":str,\"price\":number?,\"cost\":number?,\"stock\":number?,\"sold_out\":bool?,\"active\":bool?,\"description\":str?}\n"
    "7) deactivate_product (nonaktifkan produk): {\"type\":\"deactivate_product\",\"name\":str}\n"
    "8) delete_product (hapus produk): {\"type\":\"delete_product\",\"name\":str}\n"
    "9) deactivate_category (nonaktifkan kategori): {\"type\":\"deactivate_category\",\"name\":str,\"kind\":\"makanan|minuman|retail|vendor\"?}\n"
    "10) delete_category (hapus kategori): {\"type\":\"delete_category\",\"name\":str,\"kind\":\"makanan|minuman|retail|vendor\"?}\n"
    "ATURAN PENTING:\n"
    "- Jika admin MENEMPEL/menyebut BANYAK produk sekaligus (beberapa baris atau dipisah koma), WAJIB pakai SATU create_products_bulk berisi array items (JANGAN banyak blok aksi). "
    "Tebak 'kind' & 'category_name' yang masuk akal per item; default kind='retail' bila tak jelas. Jika harga tak tertera, set price 0 dan ingatkan admin melengkapi.\n"
    "- 'nonaktifkan/matikan' -> deactivate_*, 'hapus/buang' -> delete_*. Catatan: menghapus data yang sudah dipakai transaksi akan otomatis dinonaktifkan (soft delete) demi keamanan.\n"
    "- DISKON bersifat per-transaksi (diterapkan kasir saat bayar), bukan data master; untuk diskon berikan SARAN saja, jangan buat blok aksi.\n"
    "- Gunakan KONTEKS DATA untuk mencocokkan nama kategori/vendor yang sudah ada dan hindari duplikat. Jangan mengarang id. "
    "Selalu ingatkan bahwa admin akan menekan tombol 'Terapkan' untuk mengeksekusi."
)

# Panduan singkat aplikasi — dipakai Asisten AI untuk menjawab pertanyaan cara pakai fitur.
APP_GUIDE = (
    "POS Grand Aceh Kuliner (web + APK Android). Modul: POS Kasir (dine-in meja, take-away, retail, diskon, "
    "bayar cash/QRIS/debit), Shift (buka/tutup shift), Kas (setoran/pengambilan), Produk & Stok (kategori, "
    "stok, import/export Excel, opname), Vendor (bagi hasil), Laporan (harian/mingguan/bulanan, ekspor Excel/PDF, "
    "kirim WhatsApp), AI (tanya-jawab data dan usulan aksi), Pengaturan (pengguna/role, meja, perangkat printer, "
    "AI provider, installer/update dari vibecoder.co.id, diagnosa/lapor bug, versi, backup/restore, reset data). "
    "Update server: ./update-vibecoder-pi.sh atau tombol Update Sekarang (vibecoder.co.id). "
    "APK: isi alamat server http://IP-server di Pengaturan Server saat login. Fitur lapor bug: Pengaturan > Diagnostik."
)

async def _report_context(date_str):
    today = date_str or datetime.now(WIB).strftime("%Y-%m-%d")
    try:
        d = datetime.strptime(today, "%Y-%m-%d")
    except Exception:
        today = datetime.now(WIB).strftime("%Y-%m-%d")
        d = datetime.strptime(today, "%Y-%m-%d")
    yday = (d - timedelta(days=1)).strftime("%Y-%m-%d")
    s_today = await report_summary(date_str=today, admin={"role": "admin", "id": "chat"})
    s_yday = await report_summary(date_str=yday, admin={"role": "admin", "id": "chat"})
    p_items, p_total = await _purchase_summary(today)
    return {
        "tanggal": today,
        "penjualan_hari_ini": s_today,
        "penjualan_kemarin": s_yday,
        "belanja_hari_ini": {"total": p_total, "jumlah_item": len(p_items)},
    }

async def _assistant_context():
    cats = await db.categories.find({}, {"_id": 0, "name": 1, "type": 1, "active": 1}).to_list(500)
    vendors = await db.vendors.find({}, {"_id": 0, "name": 1}).sort("name", 1).to_list(500)
    pms = await db.payment_methods.find({}, {"_id": 0, "name": 1, "type": 1}).to_list(100)
    pcount = await db.products.count_documents({})
    # Data laporan (hari ini & kemarin) agar AI bisa menjawab pertanyaan penjualan juga.
    try:
        rep = await _report_context(None)
    except Exception:
        rep = {}
    return {
        "kategori": [{"nama": c.get("name"), "tipe": c.get("type")} for c in cats],
        "vendor": [v.get("name") for v in vendors],
        "metode_pembayaran": [{"nama": p.get("name"), "tipe": p.get("type")} for p in pms],
        "jumlah_produk": pcount,
        "laporan_penjualan": rep,
        "panduan_aplikasi": APP_GUIDE,
    }

def _parse_action(text):
    import json, re
    m = re.search(r"<ACTION>\s*(\{.*?\})\s*</ACTION>", text or "", re.S)
    if not m:
        return None, (text or "").strip()
    try:
        action = json.loads(m.group(1))
    except Exception:
        return None, (text or "").strip()
    clean = ((text[:m.start()] + text[m.end():]) or "").strip()
    return action, clean

def _to_num(v, default=0.0):
    if isinstance(v, (int, float)):
        return float(v)
    import re as _r
    s = _r.sub(r"[^0-9.\-]", "", str(v or ""))
    try:
        return float(s) if s not in ("", "-", ".") else float(default)
    except Exception:
        return float(default)

def _slug_sku(name):
    import re as _r
    base = _r.sub(r"[^A-Za-z0-9]+", "-", (name or "").upper()).strip("-")[:12] or "SKU"
    return f"{base}-{new_id()[:4].upper()}"

@api.post("/ai/assistant/chat")
async def assistant_chat(body: AIAssistantChatIn, admin: dict = Depends(admin_or_kasir)):
    import json
    if not (body.message or "").strip():
        raise HTTPException(400, "Pesan kosong")
    sid = body.session_id or new_id()
    sess = await db.ai_assistant_sessions.find_one({"id": sid}) or {"id": sid, "messages": []}
    history = sess.get("messages", [])
    ctx = await _assistant_context()
    role_line = "Pengguna berperan: admin (boleh menerapkan aksi)." if admin.get("role") == "admin" else "Pengguna berperan: kasir (HANYA bertanya — jangan sertakan blok aksi, perubahan data hanya admin)."
    system = role_line + "\n" + ASSISTANT_SYSTEM + "\n\nKONTEKS DATA SAAT INI:\n" + json.dumps(ctx, ensure_ascii=False)
    msgs = history + [{"role": "user", "content": body.message}]
    reply = await _ai_chat(msgs, system=system, feature="assistant", temperature=0.3, max_tokens=1500)
    action, clean = _parse_action(reply)
    history = (history + [{"role": "user", "content": body.message},
                          {"role": "assistant", "content": reply}])[-20:]
    title = (body.message or "").strip()[:60] or "Percakapan"
    await db.ai_assistant_sessions.update_one({"id": sid},
        {"$set": {"id": sid, "messages": history, "updated_at": now_utc().isoformat()},
         "$setOnInsert": {"title": title, "created_at": now_utc().isoformat()}}, upsert=True)
    return {"session_id": sid, "reply": clean or "(tidak ada balasan)", "action": action}

@api.get("/ai/assistant/sessions")
async def assistant_sessions(admin: dict = Depends(admin_or_kasir)):
    docs = await db.ai_assistant_sessions.find({}, {"_id": 0}).sort("updated_at", -1).to_list(50)
    out = []
    for d in docs:
        msgs = d.get("messages", [])
        title = d.get("title") or next((m.get("content", "") for m in msgs if m.get("role") == "user"), "Percakapan")
        out.append({"id": d["id"], "title": (title or "Percakapan")[:60],
                    "updated_at": d.get("updated_at"), "count": len(msgs)})
    return {"sessions": out}

@api.get("/ai/assistant/sessions/{sid}")
async def assistant_session_detail(sid: str, admin: dict = Depends(admin_or_kasir)):
    d = await db.ai_assistant_sessions.find_one({"id": sid}, {"_id": 0})
    if not d:
        raise HTTPException(404, "Sesi tidak ditemukan")
    msgs = []
    for m in d.get("messages", []):
        if m.get("role") == "assistant":
            _, clean = _parse_action(m.get("content", ""))
            msgs.append({"role": "assistant", "text": clean or m.get("content", "")})
        else:
            msgs.append({"role": "user", "text": m.get("content", "")})
    return {"id": sid, "messages": msgs}

@api.delete("/ai/assistant/sessions/{sid}")
async def assistant_session_delete(sid: str, admin: dict = Depends(admin_or_kasir)):
    await db.ai_assistant_sessions.delete_one({"id": sid})
    return {"deleted": True}

async def _resolve_or_create_category(name, kind):
    kind = kind if kind in ("makanan", "minuman", "retail", "vendor") else "retail"
    name = (name or "").strip() or "Umum"
    cat = await db.categories.find_one({"name": {"$regex": f"^{re.escape(name)}$", "$options": "i"}, "type": kind})
    if cat:
        return cat["id"], False
    doc = {"id": new_id(), "name": name, "type": kind, "sort_order": 0, "active": True,
           "created_at": now_utc().isoformat()}
    await db.categories.insert_one(doc)
    return doc["id"], True

async def _create_one_product(a):
    """Create a single product from an action dict. Returns a human message. Raises on validation error."""
    name = (a.get("name") or "").strip()
    if not name:
        raise HTTPException(400, "Nama produk wajib diisi")
    price = _to_num(a.get("price"), 0)
    cost = _to_num(a.get("cost"), 0)
    if price < 0 or cost < 0:
        raise HTTPException(400, f"Harga/HPP '{name}' tidak boleh negatif")
    kind = a.get("kind") if a.get("kind") in ("makanan", "minuman", "retail", "vendor") else "retail"
    cat_id, cat_created = await _resolve_or_create_category(a.get("category_name") or "Umum", kind)
    sku = (a.get("sku") or "").strip() or _slug_sku(name)
    if await db.products.find_one({"sku": sku}):
        sku = _slug_sku(name)
    vendor_id = None
    if a.get("vendor_name"):
        v = await db.vendors.find_one({"name": {"$regex": f"^{re.escape(str(a.get('vendor_name')).strip())}$", "$options": "i"}})
        vendor_id = v["id"] if v else None
    doc = {"id": new_id(), "name": name, "sku": sku, "category_id": cat_id, "type": kind,
           "price": price, "cost": cost, "vendor_id": vendor_id, "vendor_share_percent": None,
           "description": str(a.get("description") or ""), "image": "", "active": True,
           "sold_out": False, "stock": int(_to_num(a.get("stock"), 0)), "min_stock": 10,
           "track_stock": kind == "retail", "created_at": now_utc().isoformat()}
    await db.products.insert_one(doc)
    extra = " (kategori baru)" if cat_created else ""
    return f"'{name}' (SKU {sku}, Rp {int(price):,}){extra}".replace(",", ".")

async def _find_product(name):
    name = (name or "").strip()
    if not name:
        raise HTTPException(400, "Sebutkan nama/SKU produk")
    p = await db.products.find_one({"$or": [
        {"name": {"$regex": f"^{re.escape(name)}$", "$options": "i"}},
        {"sku": {"$regex": f"^{re.escape(name)}$", "$options": "i"}}]})
    if not p:
        raise HTTPException(404, f"Produk '{name}' tidak ditemukan")
    return p

async def _find_category(name, kind=None):
    name = (name or "").strip()
    if not name:
        raise HTTPException(400, "Sebutkan nama kategori")
    q = {"name": {"$regex": f"^{re.escape(name)}$", "$options": "i"}}
    if kind in ("makanan", "minuman", "retail", "vendor"):
        q["type"] = kind
    c = await db.categories.find_one(q)
    if not c:
        raise HTTPException(404, f"Kategori '{name}' tidak ditemukan")
    return c

@api.post("/ai/assistant/apply")
async def assistant_apply(body: AIAssistantApplyIn, admin: dict = Depends(require_admin)):
    a = body.action or {}
    t = a.get("type")
    if t == "create_category":
        name = (a.get("name") or "").strip()
        kind = a.get("kind") if a.get("kind") in ("makanan", "minuman", "retail", "vendor") else "retail"
        if not name:
            raise HTTPException(400, "Nama kategori wajib diisi")
        if await db.categories.find_one({"name": {"$regex": f"^{re.escape(name)}$", "$options": "i"}, "type": kind}):
            raise HTTPException(400, f"Kategori '{name}' ({kind}) sudah ada")
        doc = {"id": new_id(), "name": name, "type": kind, "sort_order": 0, "active": True,
               "created_at": now_utc().isoformat()}
        await db.categories.insert_one(doc)
        return {"ok": True, "message": f"Kategori '{name}' ({kind}) dibuat.", "entity": "category"}

    if t == "create_vendor":
        name = (a.get("name") or "").strip()
        if not name:
            raise HTTPException(400, "Nama vendor wajib diisi")
        if await db.vendors.find_one({"name": {"$regex": f"^{re.escape(name)}$", "$options": "i"}}):
            raise HTTPException(400, f"Vendor '{name}' sudah ada")
        doc = {"id": new_id(), "name": name, "contact": str(a.get("contact") or ""),
               "note": str(a.get("note") or ""), "active": True, "created_at": now_utc().isoformat()}
        await db.vendors.insert_one(doc)
        return {"ok": True, "message": f"Vendor '{name}' dibuat.", "entity": "vendor"}

    if t == "create_payment_method":
        name = (a.get("name") or "").strip()
        pm_type = a.get("pm_type") if a.get("pm_type") in ("cash", "qris", "card") else "cash"
        if not name:
            raise HTTPException(400, "Nama metode pembayaran wajib diisi")
        if await db.payment_methods.find_one({"name": {"$regex": f"^{re.escape(name)}$", "$options": "i"}}):
            raise HTTPException(400, f"Metode pembayaran '{name}' sudah ada")
        doc = {"id": new_id(), "name": name, "type": pm_type, "active": True}
        await db.payment_methods.insert_one(doc)
        return {"ok": True, "message": f"Metode pembayaran '{name}' ({pm_type}) dibuat.", "entity": "payment_method"}

    if t == "create_product":
        msg = await _create_one_product(a)
        return {"ok": True, "message": f"Produk {msg} dibuat.", "entity": "product"}

    if t == "create_products_bulk":
        items = a.get("items") or []
        if not isinstance(items, list) or not items:
            raise HTTPException(400, "Daftar produk kosong")
        if len(items) > 100:
            raise HTTPException(400, "Maksimal 100 produk per sekali proses")
        ok_msgs, errors = [], []
        for it in items:
            try:
                ok_msgs.append(await _create_one_product(it if isinstance(it, dict) else {}))
            except HTTPException as e:
                errors.append(f"{(it or {}).get('name', '?')}: {e.detail}")
            except Exception as e:
                errors.append(f"{(it or {}).get('name', '?')}: {e}")
        summary = f"{len(ok_msgs)} produk dibuat"
        if errors:
            summary += f", {len(errors)} gagal"
        return {"ok": True, "message": summary + ".", "entity": "product",
                "results": {"created": ok_msgs, "errors": errors}}

    if t == "update_product":
        name = (a.get("name") or a.get("sku") or "").strip()
        if not name:
            raise HTTPException(400, "Sebutkan nama/SKU produk yang akan diubah")
        p = await _find_product(name)
        upd = {}
        if a.get("price") is not None:
            upd["price"] = _to_num(a.get("price"))
        if a.get("cost") is not None:
            upd["cost"] = _to_num(a.get("cost"))
        if a.get("stock") is not None:
            upd["stock"] = int(_to_num(a.get("stock")))
        if a.get("sold_out") is not None:
            upd["sold_out"] = bool(a.get("sold_out"))
        if a.get("active") is not None:
            upd["active"] = bool(a.get("active"))
        if a.get("description") is not None:
            upd["description"] = str(a.get("description"))
        if upd.get("price", 0) < 0 or upd.get("cost", 0) < 0:
            raise HTTPException(400, "Harga/HPP tidak boleh negatif")
        if not upd:
            raise HTTPException(400, "Tidak ada perubahan untuk diterapkan")
        await db.products.update_one({"id": p["id"]}, {"$set": upd})
        changed = ", ".join(f"{k}={v}" for k, v in upd.items())
        return {"ok": True, "message": f"Produk '{p['name']}' diperbarui ({changed}).", "entity": "product"}

    if t == "deactivate_product":
        p = await _find_product(a.get("name") or a.get("sku"))
        await db.products.update_one({"id": p["id"]}, {"$set": {"active": False}})
        return {"ok": True, "message": f"Produk '{p['name']}' dinonaktifkan.", "entity": "product"}

    if t == "delete_product":
        p = await _find_product(a.get("name") or a.get("sku"))
        used = await db.orders.count_documents({"items.product_id": p["id"]})
        if used:
            await db.products.update_one({"id": p["id"]}, {"$set": {"active": False}})
            return {"ok": True, "message": f"Produk '{p['name']}' pernah dipakai transaksi, jadi DINONAKTIFKAN (bukan dihapus).", "entity": "product"}
        await db.products.delete_one({"id": p["id"]})
        return {"ok": True, "message": f"Produk '{p['name']}' dihapus.", "entity": "product"}

    if t == "deactivate_category":
        c = await _find_category(a.get("name"), a.get("kind"))
        await db.categories.update_one({"id": c["id"]}, {"$set": {"active": False}})
        return {"ok": True, "message": f"Kategori '{c['name']}' dinonaktifkan.", "entity": "category"}

    if t == "delete_category":
        c = await _find_category(a.get("name"), a.get("kind"))
        used = await db.products.count_documents({"category_id": c["id"]})
        if used:
            await db.categories.update_one({"id": c["id"]}, {"$set": {"active": False}})
            return {"ok": True, "message": f"Kategori '{c['name']}' dipakai {used} produk, jadi DINONAKTIFKAN (bukan dihapus).", "entity": "category"}
        await db.categories.delete_one({"id": c["id"]})
        return {"ok": True, "message": f"Kategori '{c['name']}' dihapus.", "entity": "category"}

    raise HTTPException(400, f"Jenis aksi tidak didukung: {t}")

import json as _json, re as _re

def _num(v):
    if isinstance(v, (int, float)):
        return float(v)
    s = _re.sub(r"[^0-9]", "", str(v))
    return float(s) if s else 0.0

def _extract_json_list(text):
    t = (text or "").strip()
    m = _re.search(r"\[.*\]", t, _re.S)
    if m:
        t = m.group(0)
    try:
        data = _json.loads(t)
    except Exception:
        return []
    out = []
    for d in (data if isinstance(data, list) else []):
        if not isinstance(d, dict):
            continue
        name = str(d.get("name", "")).strip()
        if name:
            out.append({"name": name, "qty": _num(d.get("qty", 1)) or 1, "unit_cost": _num(d.get("unit_cost", 0))})
    return out

class AIInvoiceIn(BaseModel):
    image: str

@api.post("/ai/parse-invoice")
async def ai_parse_invoice(body: AIInvoiceIn, admin: dict = Depends(admin_or_input)):
    cfg = await _ai_cfg("vision")
    if not (cfg["api_key"] and cfg["base_url"] and cfg["model"]):
        raise HTTPException(400, "Konfigurasi AI 'Baca Faktur (Vision)' belum lengkap di Pengaturan AI")
    img = body.image if body.image.startswith("data:") else f"data:image/jpeg;base64,{body.image}"
    system = "Anda asisten yang membaca foto faktur/nota pembelian toko."
    prompt = ('Baca foto faktur berikut dan ekstrak daftar barang. Kembalikan HANYA JSON array. '
              'Tiap elemen: {"name": "nama barang", "sku": "kode/barcode jika terlihat (else kosong)", "qty": angka, "unit_cost": angka}. '
              'unit_cost = harga beli per unit (Rupiah, angka saja tanpa titik/koma). Tanpa teks lain.')
    from openai import OpenAI

    def run():
        client = OpenAI(api_key=cfg["api_key"], base_url=cfg["base_url"])
        r = client.chat.completions.create(
            model=cfg["model"],
            messages=[
                {"role": "system", "content": system},
                {"role": "user", "content": [
                    {"type": "text", "text": prompt},
                    {"type": "image_url", "image_url": {"url": img}},
                ]},
            ],
            max_tokens=1500, temperature=0,
        )
        return r.choices[0].message.content or ""
    try:
        raw = await asyncio.to_thread(run)
    except Exception as e:
        logger.error(f"parse-invoice AI error: {e}")
        raise HTTPException(400, "Model AI yang dipilih tidak mendukung pembacaan gambar. Ganti model Vision di Pengaturan AI.")
    return {"items": _extract_json_list(raw)}

@api.post("/ai/product-description")
async def ai_description(body: AIDescIn, admin: dict = Depends(admin_or_input)):
    try:
        system = "Anda copywriter menu F&B & retail Indonesia. Tulis deskripsi produk singkat, menggugah selera, maksimal 2 kalimat, bahasa Indonesia. Jangan pakai emoji."
        prompt = f"Produk: {body.name}\nTipe: {body.type}\nKategori: {body.category}\nKata kunci: {body.keywords}\nTulis deskripsi produk."
        text = await _gemini_text(system, prompt, "description")
        return {"description": text}
    except Exception as e:
        logger.error(f"AI desc error: {e}")
        raise HTTPException(500, f"AI gagal: {e}")

@api.post("/ai/product-image")
async def ai_image(body: AIImageIn, admin: dict = Depends(admin_or_input)):
    try:
        prompt = f"Professional appetizing product photo of '{body.name}'. {body.description}. Clean studio background, top menu photography, high detail, no text overlay."
        image = await _gemini_image(prompt)
        if not image:
            raise HTTPException(500, "Tidak ada gambar dihasilkan")
        stored = await _save_image_local(image)
        return {"image": stored or image}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"AI image error: {e}")
        msg = str(e)
        if "RESOURCE_EXHAUSTED" in msg or "429" in msg:
            raise HTTPException(429, "Gambar AI belum aktif di akun Gemini Anda (kuota gambar free tier = 0). Aktifkan billing di Google Cloud/AI Studio untuk memakainya, atau unggah gambar produk secara manual.")
        raise HTTPException(500, f"AI gambar gagal: {e}")

@api.post("/reports/ai-summary")
async def ai_summary(body: AISummaryIn, admin: dict = Depends(admin_or_kasir)):
    d = body.date or now_utc().strftime("%Y-%m-%d")
    summary = await report_summary(date_str=d, admin=admin)
    try:
        system = ("Anda analis bisnis F&B & retail. Tulis LAPORAN penjualan harian dalam bahasa Indonesia yang tegas dan actionable. "
                  "Struktur: (1) Ringkasan singkat, (2) Sorotan per kategori (Makanan/Minuman/Retail), (3) 2-3 insight, (4) 1-2 rekomendasi. Tanpa emoji.")
        cr = summary.get("category_report", {})
        mk, mn, rt = cr.get("makanan", {}), cr.get("minuman", {}), cr.get("retail", {})

        def _cats(g):
            return ", ".join(f"{c['name']} Rp{c['total']:,.0f}" for c in g.get("categories", [])[:6]) or "-"
        prompt = (f"Data penjualan {d}:\n"
                  f"Total: Rp{summary['total_sales']:,.0f} dari {summary['order_count']} order. Laba kotor: Rp{summary['gross_profit']:,.0f}.\n"
                  f"Dine-in: Rp{summary['by_type']['dine_in']['total']:,.0f}; Take away: Rp{summary['by_type']['take_away']['total']:,.0f}; Retail: Rp{summary['by_type']['retail']['total']:,.0f}.\n"
                  f"Makanan Rp{mk.get('total', 0):,.0f} (rincian: {_cats(mk)}).\n"
                  f"Minuman Rp{mn.get('total', 0):,.0f} (rincian: {_cats(mn)}).\n"
                  f"Retail gabungan Rp{rt.get('total', 0):,.0f}.\n"
                  f"Total diskon: Rp{summary['total_discount']:,.0f}. Kas masuk: Rp{summary['cash_in']:,.0f}, kas keluar: Rp{summary['cash_out']:,.0f}.\n"
                  f"Produk terlaris: {', '.join(p['name'] for p in summary['top_products'][:5])}.\n"
                  f"Stok retail menipis: {len(summary.get('low_stock', []))} produk.\n"
                  f"Buat laporan analitik.")
        text = await _gemini_text(system, prompt, "summary")
        return {"date": d, "summary": text, "data": summary}
    except Exception as e:
        logger.error(f"AI summary error: {e}")
        raise HTTPException(500, f"AI ringkasan gagal: {e}")

# ================================================================== REPORT EXPORT & WHATSAPP (wacloud.id gateway)
WEBHOOK_CRON_SECRET = os.environ.get('WEBHOOK_CRON_SECRET')
WACLOUD_DEFAULT_BASE = "https://app.wacloud.id/api/v1"

async def _wa_config():
    return await db.settings.find_one({"_id": "wa"}, {"_id": 0}) or {}

async def _wa_configured():
    c = await _wa_config()
    return bool(c.get("api_key") and c.get("device_id"))

def _wa_normalize(num):
    n = "".join(ch for ch in str(num) if ch.isdigit())
    if n.startswith("0"):
        n = "62" + n[1:]
    return n

async def _wacloud_request(method, path, **kw):
    import httpx
    cfg = await _wa_config()
    api_key = cfg.get("api_key")
    base = (cfg.get("base_url") or WACLOUD_DEFAULT_BASE).rstrip("/")
    if not api_key:
        raise HTTPException(400, "API Key WhatsApp (wacloud.id) belum diisi di Pengaturan")
    headers = {"X-Api-Key": api_key, "Content-Type": "application/json"}
    async with httpx.AsyncClient(timeout=30) as c:
        return await c.request(method, f"{base}{path}", headers=headers, **kw)

async def _wa_send_text(to, text):
    cfg = await _wa_config()
    device_id = cfg.get("device_id")
    if not device_id:
        raise HTTPException(400, "Device WhatsApp belum dipilih di Pengaturan")
    return await _wacloud_request("POST", "/messages", json={
        "device_id": device_id, "to": _wa_normalize(to),
        "message_type": "text", "text": text,
    })

def _report_lines(d, s, ai_text=None):
    cr = s.get("category_report", {})
    L = ["*Laporan Grand Aceh Kuliner*", f"Tanggal: {d}", "",
         f"Total Penjualan: Rp{s['total_sales']:,.0f}",
         f"Jumlah Order: {s['order_count']}",
         f"Laba Kotor: Rp{s['gross_profit']:,.0f}",
         f"Total Diskon: Rp{s['total_discount']:,.0f}", "",
         "Per Kategori:",
         f"- Makanan: Rp{cr.get('makanan', {}).get('total', 0):,.0f}",
         f"- Minuman: Rp{cr.get('minuman', {}).get('total', 0):,.0f}",
         f"- Retail: Rp{cr.get('retail', {}).get('total', 0):,.0f}"]
    if s.get("by_payment"):
        L += ["", "Metode Bayar:"] + [f"- {k}: Rp{v:,.0f}" for k, v in s["by_payment"].items()]
    if s.get("top_products"):
        L += ["", "Produk Terlaris:"] + [f"- {p['name']} (x{p['qty']})" for p in s["top_products"][:5]]
    if ai_text:
        L += ["", "Analisis AI:", ai_text]
    return L

async def _purchase_summary(d):
    start, end = wib_day_range(d)
    items = await db.purchases.find({"created_at": {"$gte": start, "$lt": end}}, {"_id": 0}).sort("created_at", 1).to_list(2000)
    total = sum(float(x.get("total_cost", 0)) for x in items)
    return items, total

def _purchase_report_lines(d, items, total):
    L = ["*Laporan Belanja Grand Aceh Kuliner*", f"Tanggal: {d}", "",
         f"Total Belanja: Rp{total:,.0f}",
         f"Jumlah Item: {len(items)}", ""]
    if items:
        L.append("Rincian:")
        for x in items[:40]:
            L.append(f"- {x.get('product_name', '?')} x{x.get('qty', 0)} = Rp{float(x.get('total_cost', 0)):,.0f}")
    else:
        L.append("Tidak ada pembelian pada tanggal ini.")
    return L

def _report_excel(d, s):
    import openpyxl
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Laporan"
    for ln in _report_lines(d, s):
        ws.append([ln.replace("*", "")])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

def _report_pdf(d, s, ai_text=None):
    from fpdf import FPDF
    pdf = FPDF(); pdf.add_page(); pdf.set_font("Helvetica", size=12)
    for ln in _report_lines(d, s, ai_text):
        txt = ln.replace("*", "").encode("latin-1", "replace").decode("latin-1")
        pdf.multi_cell(pdf.epw, 7, txt or " ")
    return io.BytesIO(bytes(pdf.output()))

@api.get("/reports/export/excel")
async def export_report_excel(date_str: Optional[str] = Query(None, alias="date"), user: dict = Depends(get_current_user)):
    d = date_str or wib_today()
    s = await report_summary(date_str=d, admin=user)
    return StreamingResponse(_report_excel(d, s), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename=laporan-{d}.xlsx"})

@api.get("/reports/export/pdf")
async def export_report_pdf(date_str: Optional[str] = Query(None, alias="date"), user: dict = Depends(get_current_user)):
    d = date_str or wib_today()
    s = await report_summary(date_str=d, admin=user)
    return StreamingResponse(_report_pdf(d, s), media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename=laporan-{d}.pdf"})

# -------------------------------------------------------- VENDOR SETTLEMENT
async def _vendor_report(date_str=None, start=None, end=None):
    if start and end:
        s_utc, _ = wib_day_range(start)
        _, e_utc = wib_day_range(end)
        label = f"{start} s/d {end}"
    else:
        d = date_str or wib_today()
        s_utc, e_utc = wib_day_range(d)
        label = d
    q = {"status": "paid", "created_at": {"$gte": s_utc, "$lt": e_utc}}
    orders = await db.orders.find(q, {"_id": 0}).to_list(20000)
    vendors = await db.vendors.find({}, {"_id": 0}).to_list(500)
    vmap = {v["id"]: v for v in vendors}
    agg = {}
    for o in orders:
        for it in o.get("items", []):
            if it.get("type") == "vendor" and it.get("vendor_id"):
                vid = it["vendor_id"]
                a = agg.setdefault(vid, {"vendor_id": vid,
                                         "vendor_name": vmap.get(vid, {}).get("name", "(vendor dihapus)"),
                                         "qty": 0, "gross": 0, "vendor_share": 0})
                line = it["price"] * it["qty"]
                a["qty"] += it["qty"]
                a["gross"] += line
                a["vendor_share"] += it.get("vendor_total", round(line * float(it.get("vendor_share_percent") or 0) / 100, 2))
    rows = sorted(agg.values(), key=lambda x: x["gross"], reverse=True)
    for r in rows:
        r["gross"] = round(r["gross"], 2)
        r["vendor_share"] = round(r["vendor_share"], 2)
        r["outlet_share"] = round(r["gross"] - r["vendor_share"], 2)
    total_gross = round(sum(r["gross"] for r in rows), 2)
    total_vendor = round(sum(r["vendor_share"] for r in rows), 2)
    return {"label": label, "rows": rows, "total_gross": total_gross,
            "total_vendor_share": total_vendor, "total_outlet_share": round(total_gross - total_vendor, 2)}

def _vendor_report_lines(rep):
    L = ["*Laporan Bagi Hasil Vendor - Grand Aceh Kuliner*", f"Periode: {rep['label']}", "",
         f"Total Omzet Vendor: Rp{rep['total_gross']:,.0f}",
         f"Total Bagi Hasil Vendor: Rp{rep['total_vendor_share']:,.0f}",
         f"Bagian Outlet: Rp{rep['total_outlet_share']:,.0f}", ""]
    if rep["rows"]:
        L.append("Rincian per Vendor:")
        for r in rep["rows"]:
            L.append(f"- {r['vendor_name']} (x{r['qty']}): Omzet Rp{r['gross']:,.0f} -> Vendor Rp{r['vendor_share']:,.0f}")
    else:
        L.append("Belum ada penjualan produk vendor pada periode ini.")
    return L

@api.get("/reports/vendors")
async def report_vendors(date_str: Optional[str] = Query(None, alias="date"),
                         start: Optional[str] = None, end: Optional[str] = None,
                         admin: dict = Depends(admin_or_kasir)):
    return await _vendor_report(date_str, start, end)

@api.get("/reports/vendors/export/excel")
async def export_vendor_excel(date_str: Optional[str] = Query(None, alias="date"),
                              start: Optional[str] = None, end: Optional[str] = None,
                              admin: dict = Depends(admin_or_kasir)):
    import openpyxl
    rep = await _vendor_report(date_str, start, end)
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Bagi Hasil Vendor"
    ws.append(["Vendor", "Qty", "Omzet", "Bagi Hasil Vendor", "Bagian Outlet"])
    for r in rep["rows"]:
        ws.append([r["vendor_name"], r["qty"], r["gross"], r["vendor_share"], r["outlet_share"]])
    ws.append([])
    ws.append(["TOTAL", "", rep["total_gross"], rep["total_vendor_share"], rep["total_outlet_share"]])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename=bagi-hasil-vendor-{rep['label']}.xlsx"})

@api.get("/reports/vendors/export/pdf")
async def export_vendor_pdf(date_str: Optional[str] = Query(None, alias="date"),
                            start: Optional[str] = None, end: Optional[str] = None,
                            admin: dict = Depends(admin_or_kasir)):
    from fpdf import FPDF
    rep = await _vendor_report(date_str, start, end)
    pdf = FPDF(); pdf.add_page(); pdf.set_font("Helvetica", size=12)
    for ln in _vendor_report_lines(rep):
        txt = ln.replace("*", "").encode("latin-1", "replace").decode("latin-1")
        pdf.multi_cell(pdf.epw, 7, txt or " ")
    out = io.BytesIO(bytes(pdf.output()))
    return StreamingResponse(out, media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename=bagi-hasil-vendor-{rep['label']}.pdf"})

class VendorWASendIn(BaseModel):
    date: Optional[str] = None
    start: Optional[str] = None
    end: Optional[str] = None
    recipients: Optional[List[str]] = None

@api.post("/reports/vendors/send-whatsapp")
async def send_vendor_whatsapp(body: VendorWASendIn, admin: dict = Depends(require_admin)):
    rep = await _vendor_report(body.date, body.start, body.end)
    doc = await db.settings.find_one({"_id": "report"}) or {}
    recips = body.recipients or doc.get("recipients", [])
    if not recips:
        raise HTTPException(400, "Belum ada nomor WhatsApp tujuan. Atur di 'WhatsApp & Laporan'.")
    text = "\n".join(_vendor_report_lines(rep))
    result = await _send_whatsapp(recips, text)
    if not any(r.get("ok") for r in result):
        raise HTTPException(400, f"Gagal kirim WhatsApp: {result[0].get('error') if result else 'tidak diketahui'}")
    return {"sent": result}

class ReportSettingsIn(BaseModel):
    whatsapp_enabled: bool = False
    whatsapp_time: str = "22:00"
    recipients: List[str] = []
    include_ai: bool = True
    send_sales: bool = True
    send_purchases: bool = False

@api.get("/settings/report")
async def get_report_settings(admin: dict = Depends(require_admin)):
    doc = await db.settings.find_one({"_id": "report"}, {"_id": 0}) or {}
    return {
        "whatsapp_enabled": doc.get("whatsapp_enabled", False),
        "whatsapp_time": doc.get("whatsapp_time", "22:00"),
        "recipients": doc.get("recipients", []),
        "include_ai": doc.get("include_ai", True),
        "send_sales": doc.get("send_sales", True),
        "send_purchases": doc.get("send_purchases", False),
        "whatsapp_configured": await _wa_configured(),
        "last_sent_date": doc.get("last_sent_date"),
    }

@api.put("/settings/report")
async def put_report_settings(body: ReportSettingsIn, admin: dict = Depends(require_admin)):
    await db.settings.update_one({"_id": "report"}, {"$set": {
        "whatsapp_enabled": body.whatsapp_enabled, "whatsapp_time": body.whatsapp_time,
        "recipients": [r.strip() for r in body.recipients if r.strip()], "include_ai": body.include_ai,
        "send_sales": body.send_sales, "send_purchases": body.send_purchases,
    }}, upsert=True)
    return {"ok": True}

async def _send_whatsapp(recipients, text):
    out = []
    for to in recipients:
        try:
            r = await _wa_send_text(to, text)
            body = {}
            try:
                body = r.json()
            except Exception:
                pass
            if r.status_code in (200, 201) and body.get("success", True):
                out.append({"to": to, "ok": True, "id": (body.get("data") or {}).get("message_id")})
            else:
                out.append({"to": to, "ok": False, "error": body.get("error") or body.get("message") or r.text})
        except HTTPException as he:
            out.append({"to": to, "ok": False, "error": he.detail})
        except Exception as e:
            out.append({"to": to, "ok": False, "error": str(e)})
    return out

class WhatsAppSendIn(BaseModel):
    date: Optional[str] = None
    recipients: Optional[List[str]] = None

@api.post("/reports/send-whatsapp")
async def send_report_whatsapp(body: WhatsAppSendIn, admin: dict = Depends(require_admin)):
    d = body.date or wib_today()
    s = await report_summary(date_str=d, admin=admin)
    doc = await db.settings.find_one({"_id": "report"}) or {}
    recips = body.recipients or doc.get("recipients", [])
    if not recips:
        raise HTTPException(400, "Belum ada nomor WhatsApp tujuan. Atur di Pengaturan.")
    ai_text = None
    if doc.get("include_ai", True):
        try:
            r = await ai_summary(AISummaryIn(date=d), admin=admin)
            ai_text = r.get("summary")
        except Exception:
            ai_text = None
    text = "\n".join(_report_lines(d, s, ai_text))
    result = await _send_whatsapp(recips, text)
    if not any(r.get("ok") for r in result):
        raise HTTPException(400, f"Gagal kirim WhatsApp: {result[0].get('error') if result else 'tidak diketahui'}")
    return {"sent": result}

async def _run_daily_report_job():
    doc = await db.settings.find_one({"_id": "report"}) or {}
    if not doc.get("whatsapp_enabled") or not doc.get("recipients"):
        return
    if not await _wa_configured():
        return
    noww = datetime.now(WIB)
    try:
        target_hour = int(str(doc.get("whatsapp_time", "22:00")).split(":")[0])
    except Exception:
        target_hour = 22
    if noww.hour != target_hour:
        return
    today = noww.strftime("%Y-%m-%d")
    if doc.get("last_sent_date") == today:
        return
    messages = []
    if doc.get("send_sales", True):
        s = await report_summary(date_str=today, admin={"role": "admin", "id": "cron"})
        ai_text = None
        if doc.get("include_ai", True):
            try:
                r = await ai_summary(AISummaryIn(date=today), admin={"role": "admin", "id": "cron"})
                ai_text = r.get("summary")
            except Exception:
                pass
        messages.append("\n".join(_report_lines(today, s, ai_text)))
    if doc.get("send_purchases", False):
        items, total = await _purchase_summary(today)
        messages.append("\n".join(_purchase_report_lines(today, items, total)))
    if not messages:
        return
    try:
        for msg in messages:
            await _send_whatsapp(doc["recipients"], msg)
        await db.settings.update_one({"_id": "report"}, {"$set": {"last_sent_date": today}}, upsert=True)
        logger.info(f"Daily WA report sent for {today}")
    except Exception as e:
        logger.error(f"Daily WA report failed: {e}")

@api.post("/cron/daily-report")
async def cron_daily_report(request: Request, background: BackgroundTasks):
    # Cron endpoints must ack 2xx immediately; enqueue/background the actual work.
    import hmac as _hmac
    auth = request.headers.get("Authorization", "")
    token = auth[7:] if auth.startswith("Bearer ") else ""
    if not (WEBHOOK_CRON_SECRET and _hmac.compare_digest(token, WEBHOOK_CRON_SECRET)):
        raise HTTPException(401, "unauthorized")
    background.add_task(_run_daily_report_job)
    return {"ok": True}

class CronNotifyIn(BaseModel):
    to: str
    message: str

@api.post("/cron/notify")
async def cron_notify(request: Request, body: CronNotifyIn):
    import hmac as _hmac
    auth = request.headers.get("Authorization", "")
    token = auth[7:] if auth.startswith("Bearer ") else ""
    if not (WEBHOOK_CRON_SECRET and _hmac.compare_digest(token, WEBHOOK_CRON_SECRET)):
        raise HTTPException(401, "unauthorized")
    if not body.to.strip():
        raise HTTPException(400, "nomor tujuan kosong")
    res = await _send_whatsapp([body.to.strip()], body.message)
    return {"sent": res}

# ---- WhatsApp Gateway (wacloud.id) config, devices & test ----
class WAConfigIn(BaseModel):
    api_key: Optional[str] = None
    base_url: Optional[str] = None
    device_id: Optional[str] = None
    device_name: Optional[str] = None

@api.get("/whatsapp/config")
async def whatsapp_get_config(admin: dict = Depends(require_admin)):
    c = await _wa_config()
    key = c.get("api_key") or ""
    return {
        "configured": bool(c.get("api_key") and c.get("device_id")),
        "api_key_set": bool(key),
        "api_key_masked": (key[:6] + "…" + key[-4:]) if len(key) > 12 else ("•" * len(key)),
        "base_url": c.get("base_url") or WACLOUD_DEFAULT_BASE,
        "device_id": c.get("device_id", ""),
        "device_name": c.get("device_name", ""),
    }

@api.put("/whatsapp/config")
async def whatsapp_put_config(body: WAConfigIn, admin: dict = Depends(require_admin)):
    upd = {}
    if body.api_key is not None and body.api_key.strip():
        upd["api_key"] = body.api_key.strip()
    if body.base_url is not None:
        upd["base_url"] = body.base_url.strip() or WACLOUD_DEFAULT_BASE
    if body.device_id is not None:
        upd["device_id"] = body.device_id.strip()
    if body.device_name is not None:
        upd["device_name"] = body.device_name.strip()
    if upd:
        await db.settings.update_one({"_id": "wa"}, {"$set": upd}, upsert=True)
    return {"ok": True}

@api.get("/whatsapp/devices")
async def whatsapp_devices(admin: dict = Depends(require_admin)):
    try:
        r = await _wacloud_request("GET", "/devices")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"Gagal menghubungi wacloud.id: {e}")
    if r.status_code != 200:
        try:
            j = r.json()
            detail = j.get("message") or j.get("error") or r.text
        except Exception:
            detail = r.text
        raise HTTPException(r.status_code if r.status_code >= 400 else 400, f"wacloud.id: {detail}")
    data = r.json()
    devices = data.get("data", data) if isinstance(data, dict) else data
    return {"devices": devices or []}

class WATestIn(BaseModel):
    to: str
    message: Optional[str] = "Tes notifikasi Grand Aceh Kuliner POS ✅"

@api.post("/whatsapp/test")
async def whatsapp_test(body: WATestIn, admin: dict = Depends(require_admin)):
    if not body.to.strip():
        raise HTTPException(400, "Nomor tujuan wajib diisi")
    res = await _send_whatsapp([body.to.strip()], body.message or "Tes")
    if not any(x.get("ok") for x in res):
        raise HTTPException(400, f"Gagal kirim: {res[0].get('error') if res else 'tidak diketahui'}")
    return {"sent": res}

# ================================================================== MEMBERS (poin loyalitas)
class MemberIn(BaseModel):
    name: str
    phone: str = ""
    points: float = 0

@api.get("/members")
async def list_members(q: Optional[str] = None, user: dict = Depends(get_current_user)):
    query = {}
    if q:
        import re as _re
        rx = _re.compile(re.escape(q), _re.I)
        query["$or"] = [{"name": rx}, {"phone": rx}]
    return await db.members.find(query, {"_id": 0}).sort("name", 1).to_list(500)

@api.post("/members")
async def create_member(body: MemberIn, user: dict = Depends(get_current_user)):
    m = {"id": new_id(), "name": body.name.strip(), "phone": body.phone.strip(),
         "points": float(body.points or 0), "total_spend": 0.0, "created_at": now_utc().isoformat()}
    if not m["name"]:
        raise HTTPException(400, "Nama wajib diisi")
    await db.members.insert_one(m)
    m.pop("_id", None)
    return m

@api.put("/members/{mid}")
async def update_member(mid: str, body: MemberIn, user: dict = Depends(get_current_user)):
    upd = {"name": body.name.strip(), "phone": body.phone.strip(), "points": float(body.points or 0)}
    r = await db.members.update_one({"id": mid}, {"$set": upd})
    if not r.matched_count:
        raise HTTPException(404, "Member tidak ditemukan")
    return {"ok": True}

@api.delete("/members/{mid}")
async def delete_member(mid: str, admin: dict = Depends(require_admin)):
    await db.members.delete_one({"id": mid})
    return {"ok": True}

@api.get("/members/search")
async def search_member(phone: str, user: dict = Depends(get_current_user)):
    if not phone.strip():
        return {"member": None}
    m = await db.members.find_one({"phone": phone.strip()}, {"_id": 0})
    return {"member": m}

# ================================================================== PROMOS
class PromoIn(BaseModel):
    name: str
    type: Literal["percent", "happy_hour", "min_spend", "package", "bogo"]
    value: float = 0
    bonus: float = 0
    start_time: str = ""
    end_time: str = ""
    days: List[int] = []
    package_items: List[dict] = []
    active: bool = True

@api.get("/promos")
async def list_promos(user: dict = Depends(get_current_user)):
    return await db.promos.find({}, {"_id": 0}).sort("created_at", -1).to_list(200)

@api.post("/promos")
async def create_promo(body: PromoIn, admin: dict = Depends(require_admin)):
    p = {"id": new_id(), "name": body.name.strip(), "type": body.type, "value": body.value,
         "bonus": body.bonus, "start_time": body.start_time, "end_time": body.end_time,
         "days": body.days, "package_items": body.package_items, "active": body.active,
         "created_at": now_utc().isoformat()}
    if not p["name"]:
        raise HTTPException(400, "Nama promo wajib diisi")
    await db.promos.insert_one(p)
    p.pop("_id", None)
    return p

@api.put("/promos/{pid}")
async def update_promo(pid: str, body: PromoIn, admin: dict = Depends(require_admin)):
    upd = {"name": body.name.strip(), "type": body.type, "value": body.value, "bonus": body.bonus,
           "start_time": body.start_time, "end_time": body.end_time, "days": body.days,
           "package_items": body.package_items, "active": body.active}
    r = await db.promos.update_one({"id": pid}, {"$set": upd})
    if not r.matched_count:
        raise HTTPException(404, "Promo tidak ditemukan")
    return {"ok": True}

@api.delete("/promos/{pid}")
async def delete_promo(pid: str, admin: dict = Depends(require_admin)):
    await db.promos.delete_one({"id": pid})
    return {"ok": True}

# ================================================================== SPLIT / PINDAH / GABUNG MEJA
class SplitIn(BaseModel):
    items: List[OrderItem]

@api.post("/orders/{oid}/split")
async def split_order(oid: str, body: SplitIn, admin: dict = Depends(require_admin)):
    if not body.items:
        raise HTTPException(400, "Pilih item yang akan dipindah")
    o = await db.orders.find_one({"id": oid})
    if not o or o["status"] != "open":
        raise HTTPException(400, "Order tidak ditemukan / bukan open bill")
    moved_ids = {it.product_id: it.qty for it in body.items}
    keep, moved = [], []
    for it in o["items"]:
        need = moved_ids.get(it["product_id"], 0)
        if need >= it["qty"]:
            moved.append({**it})
            moved_ids[it["product_id"]] = need - it["qty"]
        elif need > 0:
            moved.append({**it, "qty": need})
            keep.append({**it, "qty": it["qty"] - need})
            moved_ids[it["product_id"]] = 0
        else:
            keep.append({**it})
    if not moved:
        raise HTTPException(400, "Tidak ada item yang valid untuk dipindah")
    st_k, d_k, _ = compute_totals(keep, o["discount_type"], o["discount_value"])
    pr_k, pn_k = await _apply_promos(keep, st_k)
    t_k = max(0.0, round(st_k - d_k - pr_k, 2))
    new_doc = {
        "id": new_id(), "order_number": await gen_order_number(),
        "order_type": o["order_type"], "table_id": o.get("table_id"), "items": moved,
        "subtotal": sum(i["price"] * i["qty"] for i in moved),
        "discount_type": "none", "discount_value": 0, "discount": 0,
        "promo_discount": 0, "promos_applied": [], "redeem_discount": 0,
        "total": sum(i["price"] * i["qty"] for i in moved),
        "note": f"Split dari {o['order_number']}", "status": "open",
        "cashier_id": admin["id"], "cashier_name": admin["name"],
        "created_at": now_utc().isoformat(), "parent_order": o["id"],
    }
    await db.orders.insert_one(new_doc)
    await db.orders.update_one({"id": oid}, {"$set": {"items": keep, "subtotal": st_k,
                                                      "discount": d_k, "promo_discount": pr_k,
                                                      "promos_applied": pn_k, "total": t_k}})
    new_doc.pop("_id", None)
    return {"original": await db.orders.find_one({"id": oid}, {"_id": 0}), "new_order": new_doc}

class TableMoveIn(BaseModel):
    table_id: str

@api.patch("/orders/{oid}/table")
async def move_order_table(oid: str, body: TableMoveIn, user: dict = Depends(get_current_user)):
    o = await db.orders.find_one({"id": oid})
    if not o or o["status"] != "open":
        raise HTTPException(400, "Order tidak ditemukan / bukan open bill")
    await db.orders.update_one({"id": oid}, {"$set": {"table_id": body.table_id}})
    return {"ok": True}

class MergeIn(BaseModel):
    target_id: str

@api.post("/orders/{oid}/merge")
async def merge_orders(oid: str, body: MergeIn, admin: dict = Depends(require_admin)):
    o = await db.orders.find_one({"id": oid})
    t = await db.orders.find_one({"id": body.target_id})
    if not o or not t or o["status"] != "open" or t["status"] != "open":
        raise HTTPException(400, "Kedua order harus open bill")
    if o["id"] == t["id"]:
        raise HTTPException(400, "Target tidak boleh sama")
    merged_items = t["items"] + o["items"]
    st, d, _ = compute_totals(merged_items, t["discount_type"], t["discount_value"])
    pr, pn = await _apply_promos(merged_items, st)
    tot = max(0.0, round(st - d - pr, 2))
    await db.orders.update_one({"id": t["id"]}, {"$set": {"items": merged_items, "subtotal": st,
                                                          "discount": d, "promo_discount": pr,
                                                          "promos_applied": pn, "total": tot}})
    await db.orders.update_one({"id": oid}, {"$set": {"status": "merged", "merged_into": t["id"],
                                                      "void_reason": f"Digabung ke {t['order_number']}"}})
    return {"ok": True, "target": await db.orders.find_one({"id": t["id"]}, {"_id": 0})}

# ================================================================== LABA KOTOR PER PRODUK
@api.get("/reports/profit")
async def report_profit(start: str, end: str, admin: dict = Depends(admin_or_kasir)):
    s, e = wib_day_range(start[:10]), wib_day_range(end[:10])
    q = {"status": "paid", "created_at": {"$gte": s[0], "$lte": e[1]}}
    orders = await db.orders.find(q, {"_id": 0}).to_list(5000)
    rows = {}
    for o in orders:
        for it in o.get("items", []):
            pid = it["product_id"]
            r = rows.setdefault(pid, {"name": it["name"], "qty": 0, "revenue": 0.0, "cost": 0.0})
            r["qty"] += it["qty"]
            r["revenue"] += it["price"] * it["qty"]
            r["cost"] += (it.get("cost") or 0) * it["qty"]
    out = []
    for pid, r in rows.items():
        out.append({"product_id": pid, "name": r["name"], "qty": r["qty"],
                    "revenue": round(r["revenue"], 2), "cost": round(r["cost"], 2),
                    "profit": round(r["revenue"] - r["cost"], 2),
                    "margin": round((r["revenue"] - r["cost"]) / r["revenue"] * 100, 1) if r["revenue"] else 0})
    out.sort(key=lambda x: x["profit"], reverse=True)
    return {"rows": out, "total_revenue": round(sum(r["revenue"] for r in rows.values()), 2),
            "total_cost": round(sum(r["cost"] for r in rows.values()), 2),
            "total_profit": round(sum(r["revenue"] - r["cost"] for r in rows.values()), 2)}

# ================================================================== REKOMENDASI PEMBELIAN STOK (AI)
@api.post("/ai/purchase-recommendation")
async def purchase_recommendation(admin: dict = Depends(admin_or_kasir)):
    import json
    end = datetime.now(WIB)
    start = end - timedelta(days=30)
    days = 30
    sold = {}
    for o in await db.orders.find({"status": "paid", "created_at": {"$gte": start.isoformat()}}, {"_id": 0, "items": 1}).to_list(5000):
        for it in o.get("items", []):
            if it.get("type") == "retail":
                sold[it["product_id"]] = sold.get(it["product_id"], 0) + it["qty"]
    recs = []
    for p in await db.products.find({"type": "retail", "track_stock": True}, {"_id": 0}).to_list(2000):
        qty30 = sold.get(p["id"], 0)
        avg = qty30 / days
        stock = float(p.get("stock") or 0)
        min_stock = float(p.get("min_stock") or 10)
        suggest = max(0.0, (avg * 14) - stock)
        recs.append({"product_id": p["id"], "name": p["name"], "stock": stock,
                     "min_stock": min_stock, "sold_30d": qty30, "daily_avg": round(avg, 2),
                     "suggest": math.ceil(suggest)})
    recs.sort(key=lambda x: x["sold_30d"], reverse=True)
    ai_text = ""
    try:
        top = recs[:12]
        prompt = ("Buat ringkasan rekomendasi pembelian stok dalam 3-5 baris Bahasa Indonesia. "
                  f"Data (produk, stok, terjual 30 hari, saran beli):\n{json.dumps(top, ensure_ascii=False, default=str)}")
        ai_text = await _gemini_text("Anda analis stok restoran.", prompt, feature="summary")
    except Exception:
        ai_text = ""
    return {"rows": recs, "ai_summary": ai_text}

# ================================================================== SETTINGS OUTLET (nama/alamat/logo)
class OutletIn(BaseModel):
    name: str = ""
    address: str = ""
    phone: str = ""

@api.get("/settings/outlet")
async def get_outlet(admin: dict = Depends(require_admin)):
    doc = await db.settings.find_one({"_id": "outlet"}, {"_id": 0}) or {}
    return {"name": doc.get("name", ""), "address": doc.get("address", ""),
            "phone": doc.get("phone", ""), "logo_url": doc.get("logo_url", "")}

@api.put("/settings/outlet")
async def put_outlet(body: OutletIn, admin: dict = Depends(require_admin)):
    await db.settings.update_one({"_id": "outlet"}, {"$set": {
        "name": body.name.strip(), "address": body.address.strip(), "phone": body.phone.strip()}}, upsert=True)
    return {"ok": True}

@api.post("/settings/outlet/logo")
async def upload_logo(file: UploadFile = File(...), admin: dict = Depends(require_admin)):
    import mimetypes
    ok_types = {"image/png": ".png", "image/jpeg": ".jpg", "image/webp": ".webp", "image/gif": ".gif"}
    ext = ok_types.get(file.content_type or "")
    if not ext:
        raise HTTPException(400, "Format logo harus PNG/JPG/WEBP/GIF")
    data = await file.read()
    if len(data) > 2_000_000:
        raise HTTPException(400, "Logo maksimal 2MB")
    fname = f"logo-{new_id()}{ext}"
    (UPLOAD_DIR / fname).write_bytes(data)
    url = f"/uploads/{fname}"
    await db.settings.update_one({"_id": "outlet"}, {"$set": {"logo_url": url}}, upsert=True)
    return {"ok": True, "url": url}

# ================================================================== EXCEL
IMPORT_COLUMNS = ["nama_produk", "sku", "kategori", "tipe_produk", "harga", "harga_beli", "status_aktif", "sold_out", "deskripsi", "stok_awal"]

@api.get("/products/template")
async def download_template(user: dict = Depends(get_current_user)):
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    cats = await db.categories.find({}, {"_id": 0, "name": 1, "type": 1}).to_list(500)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produk"
    ws.append(IMPORT_COLUMNS)
    # Contoh memakai KATEGORI NYATA dari database supaya langsung valid
    ex = []
    by_type = {}
    for c in cats:
        by_type.setdefault(c["type"], []).append(c["name"])
    if by_type.get("makanan"):
        ex.append(["Nasi Goreng Aceh", "FD-001", by_type["makanan"][0], "makanan", 25000, 12000, "aktif", "tidak", "Nasi goreng khas Aceh", 0])
    if by_type.get("minuman"):
        ex.append(["Kopi Sanger", "DR-001", by_type["minuman"][0], "minuman", 15000, 6000, "aktif", "tidak", "Kopi susu khas Aceh", 0])
    if by_type.get("retail"):
        ex.append(["Keripik Pisang", "RT-001", by_type["retail"][0], "retail", 12000, 8000, "aktif", "tidak", "Keripik pisang kemasan", 50])
    if not ex:
        ex = [["Nasi Goreng Aceh", "FD-001", "", "makanan", 25000, 12000, "aktif", "tidak", "Nasi goreng khas Aceh", 0]]
    for row in ex:
        ws.append(row)
    # Styling header
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="E63946")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 13
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 30
    ws.column_dimensions["J"].width = 10

    # Lembar Petunjuk
    guide = wb.create_sheet("Petunjuk")
    guide.column_dimensions["A"].width = 26
    guide.column_dimensions["B"].width = 60
    guide.append(["Kolom", "Keterangan & Nilai yang Diterima"])
    guide.append(["nama_produk", "Nama produk (wajib)"])
    guide.append(["sku", "Kode SKU unik (wajib)"])
    guide.append(["kategori", "Nama kategori — harus SAMA dengan kategori yang ada di aplikasi. Kategori valid:"])
    for c in cats:
        guide.append(["", f"- {c['name']}  (tipe: {c['type']})"])
    if not cats:
        guide.append(["", "(belum ada kategori — buat dulu di menu Produk & Stok > Kategori)"])
    guide.append(["tipe_produk", "makanan | minuman | retail (harus sesuai tipe kategori)"])
    guide.append(["harga", "Angka (rupiah), contoh: 25000"])
    guide.append(["harga_beli", "Angka HPP/modal (0 untuk produk tanpa HPP)"])
    guide.append(["status_aktif", "aktif | nonaktif (juga diterima: ya/tidak/true/false/1/0)"])
    guide.append(["sold_out", "tidak | ya (tandai produk habis)"])
    guide.append(["deskripsi", "Teks deskripsi (opsional)"])
    guide.append(["stok_awal", "Angka stok (hanya dipakai untuk produk retail)"])
    guide.append([])
    guide.append(["CATATAN", "Hapus baris contoh sebelum mengisi data Anda. Baris yang error tidak akan diimpor; perbaiki lalu ulangi."])
    for cell in guide[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F46E5")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=template_produk_gak.xlsx"})

@api.get("/products/export")
async def export_products(admin: dict = Depends(require_admin)):
    import openpyxl
    cats = {c["id"]: c["name"] for c in await db.categories.find({}, {"_id": 0}).to_list(500)}
    products = await db.products.find({}, {"_id": 0}).sort("name", 1).to_list(5000)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produk"
    ws.append(IMPORT_COLUMNS)
    for p in products:
        ws.append([p["name"], p["sku"], cats.get(p["category_id"], ""), p["type"], p["price"], p.get("cost", 0),
                   "aktif" if p.get("active") else "nonaktif", "ya" if p.get("sold_out") else "tidak",
                   p.get("description", ""), p.get("stock", 0)])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=produk_gak.xlsx"})

async def _parse_import(file_bytes):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], ["File kosong"]
    header = [str(h).strip().lower() if h else "" for h in rows[0]]
    dict_rows = [{header[i]: raw[i] if i < len(raw) else None for i in range(len(header))} for raw in rows[1:]]
    return await _validate_import_rows(dict_rows)

async def _validate_import_rows(dict_rows):
    """Validasi baris import (dari file Excel atau JSON hasil perbaikan) -> parsed."""
    cats = {c["name"].strip().lower(): c for c in await db.categories.find({}, {"_id": 0}).to_list(500)}
    existing_skus = {p["sku"] for p in await db.products.find({}, {"_id": 0, "sku": 1}).to_list(5000)}
    parsed = []
    seen_skus = set()
    for idx, row in enumerate(dict_rows, start=2):
        errors = []
        name = str(row.get("nama_produk") or "").strip()
        sku = str(row.get("sku") or "").strip()
        cat_name = str(row.get("kategori") or "").strip()
        ptype = str(row.get("tipe_produk") or "").strip().lower()
        if not name:
            errors.append("nama produk kosong")
        if not sku:
            errors.append("SKU kosong")
        elif sku in seen_skus:
            errors.append(f"SKU '{sku}' duplikat di file")
        seen_skus.add(sku)
        cat = cats.get(cat_name.lower())
        if not cat:
            errors.append(f"kategori '{cat_name}' tidak valid")
        if ptype not in PRODUCT_TYPES:
            errors.append(f"tipe '{ptype}' tidak valid (makanan/minuman/retail)")
        elif cat and cat["type"] != ptype:
            errors.append(f"tipe produk '{ptype}' tidak sesuai tipe kategori '{cat['type']}'")
        try:
            price = float(row.get("harga") or 0)
            if price < 0:
                errors.append("harga negatif")
        except (ValueError, TypeError):
            price = 0
            errors.append("harga bukan angka")
        try:
            cost = float(row.get("harga_beli") or 0)
            if cost < 0:
                cost = 0
        except (ValueError, TypeError):
            cost = 0
        try:
            stock = int(float(row.get("stok_awal") or 0))
        except (ValueError, TypeError):
            stock = 0
        active = str(row.get("status_aktif") or "aktif").strip().lower() in ("aktif", "aktif ", "ya", "true", "1", "active")
        sold_out = str(row.get("sold_out") or "tidak").strip().lower() in ("ya", "true", "1", "sold out", "soldout")
        parsed.append({
            "row": idx, "name": name, "sku": sku, "category_id": cat["id"] if cat else None,
            "category_name": cat_name, "type": ptype, "price": price, "cost": cost, "stock": stock,
            "description": str(row.get("deskripsi") or ""), "active": active, "sold_out": sold_out,
            "exists": sku in existing_skus, "errors": errors, "valid": len(errors) == 0,
        })
    return parsed, []

@api.post("/products/import/preview")
async def import_preview(file: UploadFile = File(...), admin: dict = Depends(admin_or_input)):
    content = await file.read()
    parsed, file_errors = await _parse_import(content)
    if file_errors:
        raise HTTPException(400, file_errors[0])
    valid = [p for p in parsed if p["valid"]]
    return {"rows": parsed, "total": len(parsed), "valid_count": len(valid),
            "error_count": len(parsed) - len(valid),
            "new_count": len([p for p in valid if not p["exists"]]),
            "update_count": len([p for p in valid if p["exists"]])}

@api.post("/products/import/commit")
async def import_commit(file: UploadFile = File(...), admin: dict = Depends(admin_or_input)):
    content = await file.read()
    parsed, file_errors = await _parse_import(content)
    if file_errors:
        raise HTTPException(400, file_errors[0])
    created = updated = 0
    for p in parsed:
        if not p["valid"]:
            continue
        doc = {"name": p["name"], "sku": p["sku"], "category_id": p["category_id"], "type": p["type"],
               "price": p["price"], "cost": p["cost"], "description": p["description"], "active": p["active"],
               "sold_out": p["sold_out"], "stock": p["stock"], "track_stock": p["type"] == "retail",
               "image": ""}
        if p["exists"]:
            await db.products.update_one({"sku": p["sku"]}, {"$set": doc})
            updated += 1
        else:
            doc.update({"id": new_id(), "created_at": now_utc().isoformat()})
            await db.products.insert_one(doc)
            created += 1
    log = {"id": new_id(), "filename": file.filename, "at": now_utc().isoformat(), "by": admin["name"],
           "created": created, "updated": updated, "errors": len([p for p in parsed if not p["valid"]])}
    await db.import_logs.insert_one(log)
    log.pop("_id", None)
    return log

class ImportRowsFixIn(BaseModel):
    rows: List[dict]

@api.post("/products/import/commit-fix")
async def import_commit_fix(body: ImportRowsFixIn, admin: dict = Depends(admin_or_input)):
    """Commit hasil PERBAIKAN import (JSON) — baris yang diedit user di UI.
    Validasi ulang server-side lalu simpan yang valid."""
    if not body.rows:
        raise HTTPException(400, "Tidak ada baris untuk diimpor")
    parsed, file_errors = await _validate_import_rows(body.rows)
    created = updated = 0
    for p in parsed:
        if not p["valid"]:
            continue
        doc = {"name": p["name"], "sku": p["sku"], "category_id": p["category_id"], "type": p["type"],
               "price": p["price"], "cost": p["cost"], "description": p["description"], "active": p["active"],
               "sold_out": p["sold_out"], "stock": p["stock"], "track_stock": p["type"] == "retail",
               "image": ""}
        if p["exists"]:
            await db.products.update_one({"sku": p["sku"]}, {"$set": doc})
            updated += 1
        else:
            doc.update({"id": new_id(), "created_at": now_utc().isoformat()})
            await db.products.insert_one(doc)
            created += 1
    log = {"id": new_id(), "filename": "perbaikan-import", "at": now_utc().isoformat(), "by": admin["name"],
           "created": created, "updated": updated, "errors": len([p for p in parsed if not p["valid"]])}
    await db.import_logs.insert_one(log)
    log.pop("_id", None)
    return log

@api.get("/import-logs")
async def import_logs(admin: dict = Depends(require_admin)):
    return await db.import_logs.find({}, {"_id": 0}).sort("at", -1).to_list(100)

@api.post("/ai/assistant/import-excel")
async def assistant_import_excel(file: UploadFile = File(...), admin: dict = Depends(admin_or_input)):
    """Parse file Excel yang diupload di chat Asisten AI — return baris siap
    commit (commit-fix). Tidak menyimpan apa pun sampai admin klik Terapkan."""
    content = await file.read()
    if len(content) > 5_000_000:
        raise HTTPException(400, "File terlalu besar (maks 5MB)")
    parsed, file_errors = await _parse_import(content)
    if file_errors:
        raise HTTPException(400, file_errors[0])
    rows = [{
        "nama_produk": p["name"], "sku": p["sku"], "kategori": p.get("category_name") or "",
        "tipe_produk": p["type"], "harga": p["price"], "harga_beli": p["cost"],
        "exists": p["exists"], "errors": p["errors"], "valid": p["valid"],
    } for p in parsed]
    return {
        "rows": rows, "total": len(rows),
        "valid_count": sum(1 for r in rows if r["valid"]),
        "error_count": sum(1 for r in rows if not r["valid"]),
        "new_count": sum(1 for r in rows if r["valid"] and not r["exists"]),
        "update_count": sum(1 for r in rows if r["valid"] and r["exists"]),
    }

# ------------------------------------------------------------------ startup
@app.on_event("startup")
async def startup():
    await db.users.create_index("email", unique=True)
    await db.products.create_index("sku", unique=True)
    await db.orders.create_index("client_ref", sparse=True)
    await db.orders.create_index([("order_type", 1), ("status", 1)])
    await db.orders.create_index([("created_at", -1)])
    await db.products.create_index([("type", 1), ("active", 1)])
    await db.products.create_index([("category_id", 1)])
    try:
        await db.orders.create_index("order_number", unique=True)
    except Exception as e:
        logger.warning(f"order_number unique index not applied (resolve duplicates): {e}")
    admin_email = os.environ["ADMIN_EMAIL"].lower()
    admin_pw = os.environ["ADMIN_PASSWORD"]
    existing = await db.users.find_one({"email": admin_email})
    if not existing:
        await db.users.insert_one({"id": new_id(), "name": os.environ.get("ADMIN_NAME", "Admin"),
                                   "email": admin_email, "password_hash": hash_password(admin_pw),
                                   "role": "admin", "active": True, "created_at": now_utc().isoformat()})
    elif not verify_password(admin_pw, existing["password_hash"]):
        await db.users.update_one({"email": admin_email}, {"$set": {"password_hash": hash_password(admin_pw)}})
    # seed a cashier
    if not await db.users.find_one({"email": "kasir@grandaceh.com"}):
        await db.users.insert_one({"id": new_id(), "name": "Kasir Satu", "email": "kasir@grandaceh.com",
                                   "password_hash": hash_password("kasir123"), "role": "kasir",
                                   "active": True, "created_at": now_utc().isoformat()})
    # seed payment methods
    if await db.payment_methods.count_documents({}) == 0:
        for n, t in [("Cash", "cash"), ("QRIS", "qris"), ("Kartu Debit/Kredit", "card")]:
            await db.payment_methods.insert_one({"id": new_id(), "name": n, "type": t, "active": True})
    logger.info("Startup seeding complete")

    # Scheduler internal untuk laporan WhatsApp harian (self-hosted; tanpa cron eksternal).
    import asyncio as _asyncio

    async def _report_scheduler():
        while True:
            try:
                await _run_daily_report_job()
            except Exception as e:
                logger.error(f"report scheduler tick failed: {e}")
            await _asyncio.sleep(600)  # cek tiap 10 menit
    _asyncio.create_task(_report_scheduler())

@app.on_event("shutdown")
async def shutdown():
    client.close()

app.include_router(api)
app.add_middleware(
    CORSMiddleware,
    allow_credentials=False,
    allow_origins=[o.strip() for o in os.environ.get('CORS_ORIGINS', '*').split(',') if o.strip()],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Private Network Access (PNA): Chrome/WebView memblokir panggilan cross-origin ke
# IP privat (LAN 192.168.x / tailnet 100.x) bila respons tidak menyertakan
# Access-Control-Allow-Private-Network: true. Middleware ini dipasang PALING LUAR
# (setelah CORSMiddleware) sehingga menambah header ke SEMUA respons, termasuk
# preflight OPTIONS yang ditangani CORSMiddleware.
from starlette.middleware.base import BaseHTTPMiddleware

class PrivateNetworkAccessMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request, call_next):
        response = await call_next(request)
        try:
            response.headers["Access-Control-Allow-Private-Network"] = "true"
        except Exception:
            pass
        return response

app.add_middleware(PrivateNetworkAccessMiddleware)
